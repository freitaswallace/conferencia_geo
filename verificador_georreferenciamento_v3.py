#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verificador de Consistência de Documentos de Georreferenciamento
Aplicação GUI para cartórios - Análise multimodal com Gemini AI
Autor: Sistema Automatizado
Versão: 3.0 - Com extração para Excel integrada
"""

import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from tkinter import ttk
from pathlib import Path
import threading
from typing import List, Optional, Dict
import json
import tempfile

try:
    from pdf2image import convert_from_path
    from PIL import Image, ImageTk
    import google.generativeai as genai
    from openpyxl import load_workbook
    # Importar funções de extração do script existente
    from process_memorial_descritivo_v2 import (
        extract_table_from_pdf,
        extrair_memorial_incra,
        create_excel_file
    )
except ImportError as e:
    print(f"❌ Erro: Biblioteca necessária não encontrada: {e}")
    print("\nInstale as dependências com:")
    print("pip install pdf2image Pillow google-generativeai openpyxl --break-system-packages")
    print("\nNota: Também é necessário ter o 'poppler-utils' instalado no sistema.")
    print("Certifique-se de que process_memorial_descritivo_v2.py está no mesmo diretório.")
    sys.exit(1)


class VerificadorGeorreferenciamento:
    """Classe principal da aplicação de verificação de documentos."""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Verificador de Consistência - Georreferenciamento")
        self.root.geometry("1200x900")
        
        # Configurar fonte padrão maior para melhor legibilidade
        self.root.option_add("*Font", "Arial 12")
        
        # Variáveis para armazenar caminhos dos arquivos
        self.incra_path = tk.StringVar()
        self.projeto_path = tk.StringVar()
        self.api_key = tk.StringVar()

        # Variáveis para armazenar imagens processadas (para comparação visual)
        self.incra_images: List[Image.Image] = []
        self.projeto_images: List[Image.Image] = []

        # Variáveis para armazenar dados extraídos (nova funcionalidade v3)
        self.incra_excel_path: Optional[str] = None
        self.projeto_excel_path: Optional[str] = None
        self.incra_data: Optional[Dict] = None
        self.projeto_data: Optional[Dict] = None
        
        self._criar_interface()
        
    def _criar_interface(self):
        """Cria todos os elementos da interface gráfica."""
        
        # Frame principal com padding
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configurar grid para expansão
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # ===== SEÇÃO: API KEY =====
        ttk.Label(main_frame, text="🔑 API Key do Gemini:", 
                 font=('Arial', 14, 'bold')).grid(row=0, column=0, sticky=tk.W, pady=(0, 10))
        
        api_entry = ttk.Entry(main_frame, textvariable=self.api_key, width=40, show="*", font=('Arial', 12))
        api_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=(0, 10), padx=(10, 0))
        
        # ===== SEÇÃO: SELEÇÃO DE ARQUIVOS =====
        ttk.Separator(main_frame, orient='horizontal').grid(
            row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        
        ttk.Label(main_frame, text="📄 Documentos:", 
                 font=('Arial', 14, 'bold')).grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=(5, 0))
        
        # INCRA
        self._criar_linha_arquivo(main_frame, 3, "INCRA:", self.incra_path)

        # Projeto/Planta
        self._criar_linha_arquivo(main_frame, 4, "Projeto/Planta:", self.projeto_path)
        
        # ===== SEÇÃO: BOTÕES DE AÇÃO =====
        ttk.Separator(main_frame, orient='horizontal').grid(
            row=5, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)

        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=6, column=0, columnspan=2, pady=15)

        # Estilo para botões maiores
        style = ttk.Style()
        style.configure('Large.TButton', font=('Arial', 12, 'bold'), padding=10)

        # Botão único: INCRA vs. Projeto
        self.btn_comparar = ttk.Button(
            button_frame,
            text="📐  COMPARAR: INCRA vs. Projeto",
            command=self._comparar_projeto,
            style='Large.TButton',
            width=40
        )
        self.btn_comparar.pack(pady=5)

        # Comparação Visual Manual
        self.btn_comparacao_manual = ttk.Button(
            button_frame,
            text="👁️  Comparação Visual Manual",
            command=self._abrir_comparacao_manual,
            style='Large.TButton',
            width=40
        )
        self.btn_comparacao_manual.pack(pady=5)
        
        # ===== SEÇÃO: ÁREA DE RESULTADOS =====
        ttk.Separator(main_frame, orient='horizontal').grid(
            row=7, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        
        ttk.Label(main_frame, text="📋 Relatório de Comparação:",
                 font=('Arial', 14, 'bold')).grid(row=8, column=0, columnspan=2, sticky=tk.W, pady=(5, 0))

        # Frame para área de texto com barra de rolagem
        text_frame = ttk.Frame(main_frame)
        text_frame.grid(row=9, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)

        # Área de texto com scroll e fonte maior
        self.resultado_text = scrolledtext.ScrolledText(
            text_frame,
            width=85,
            height=22,
            wrap=tk.WORD,
            font=('Consolas', 11),
            bg='#ffffff',
            fg='#000000'
        )
        self.resultado_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Botão para salvar HTML
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=10, column=0, columnspan=2, pady=(5, 0))
        
        self.btn_salvar_html = ttk.Button(
            btn_frame,
            text="💾 Salvar Relatório em HTML",
            command=self._salvar_relatorio_html,
            state='disabled'
        )
        self.btn_salvar_html.pack(side=tk.LEFT, padx=5)
        
        # Configurar expansão da área de texto
        main_frame.rowconfigure(9, weight=1)

        # Barra de status com fonte maior
        self.status_label = ttk.Label(main_frame, text="✅ Sistema Pronto para Uso",
                                      relief=tk.SUNKEN, anchor=tk.W, font=('Arial', 11))
        self.status_label.grid(row=11, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        
        # Variável para armazenar o HTML do último relatório
        self.ultimo_relatorio_html = ""
        
    def _criar_linha_arquivo(self, parent, row, label_text, text_var):
        """Cria uma linha com label, entry e botão para seleção de arquivo."""
        ttk.Label(parent, text=label_text, font=('Arial', 13)).grid(row=row, column=0, sticky=tk.W, pady=8)
        
        entry_frame = ttk.Frame(parent)
        entry_frame.grid(row=row, column=1, sticky=(tk.W, tk.E), pady=8, padx=(10, 0))
        entry_frame.columnconfigure(0, weight=1)
        
        entry = ttk.Entry(entry_frame, textvariable=text_var, font=('Arial', 11))
        entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))
        
        btn = ttk.Button(entry_frame, text="📁 Selecionar", 
                        command=lambda: self._selecionar_arquivo(text_var))
        btn.configure(width=15)
        btn.grid(row=0, column=1)
        
    def _selecionar_arquivo(self, text_var):
        """Abre diálogo para seleção de arquivo PDF."""
        filename = filedialog.askopenfilename(
            title="Selecione o arquivo PDF",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")]
        )
        if filename:
            text_var.set(filename)
            
    def _salvar_relatorio_html(self):
        """Salva o relatório atual em arquivo HTML."""
        if not self.ultimo_relatorio_html:
            messagebox.showwarning("Aviso", "Nenhum relatório para salvar. Execute uma análise primeiro.")
            return
            
        # Abrir diálogo para salvar arquivo
        filename = filedialog.asksaveasfilename(
            title="Salvar Relatório",
            defaultextension=".html",
            filetypes=[("HTML Files", "*.html"), ("All Files", "*.*")]
        )
        
        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(self.ultimo_relatorio_html)
                messagebox.showinfo("Sucesso", f"Relatório salvo em:\n{filename}")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar arquivo:\n{str(e)}")
    
    def _abrir_comparacao_manual(self):
        """Abre janela de comparação visual manual dos documentos."""
        # Verificar se há documentos carregados
        if not self.incra_path.get():
            messagebox.showwarning(
                "Aviso", 
                "Por favor, selecione pelo menos o arquivo INCRA."
            )
            return
        
        if not self.memorial_path.get() and not self.projeto_path.get():
            messagebox.showwarning(
                "Aviso",
                "Por favor, selecione pelo menos o Memorial ou o Projeto para comparar."
            )
            return
        
        # Criar e abrir janela de comparação
        try:
            janela_comparacao = JanelaComparacaoManual(
                self.root,
                self.incra_path.get(),
                self.memorial_path.get() if self.memorial_path.get() else None,
                self.projeto_path.get() if self.projeto_path.get() else None
            )
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir comparação manual:\n{str(e)}")
    
    def _validar_entrada(self) -> bool:
        """Valida se todos os campos necessários foram preenchidos."""
        if not self.api_key.get().strip():
            messagebox.showerror("Erro", "Por favor, insira a API Key do Gemini.")
            return False

        if not self.incra_path.get():
            messagebox.showerror("Erro", "Por favor, selecione o arquivo INCRA.")
            return False

        if not self.projeto_path.get():
            messagebox.showerror("Erro", "Por favor, selecione o arquivo Projeto/Planta.")
            return False

        return True
        
    def _atualizar_status(self, mensagem: str):
        """Atualiza a barra de status."""
        self.status_label.config(text=mensagem)
        self.root.update_idletasks()
        
    def _desabilitar_botoes(self):
        """Desabilita os botões durante o processamento."""
        self.btn_comparar.config(state='disabled')

    def _habilitar_botoes(self):
        """Reabilita os botões após o processamento."""
        self.btn_comparar.config(state='normal')

    # ========== NOVAS FUNÇÕES V3: EXTRAÇÃO PARA EXCEL ==========

    def _extrair_pdf_para_excel(self, pdf_path: str, tipo: str = "normal") -> tuple[str, Dict]:
        """
        Extrai dados de um PDF memorial para Excel usando Gemini API.

        Args:
            pdf_path: Caminho do arquivo PDF
            tipo: "incra" para usar extração especializada INCRA, "normal" para outros

        Returns:
            Tupla (caminho_excel, dados_dict)
        """
        try:
            api_key = self.api_key.get().strip()

            # Criar diretório temporário para Excel se não existir
            # Usa tempfile.gettempdir() que é multiplataforma (Windows/Linux/Mac)
            output_dir = Path(tempfile.gettempdir()) / "conferencia_geo"

            # Criar diretório com permissões adequadas
            output_dir.mkdir(parents=True, exist_ok=True)

            # Verificar se o diretório foi criado
            if not output_dir.exists():
                raise RuntimeError(f"Não foi possível criar o diretório: {output_dir}")

            # Definir nome do arquivo Excel
            pdf_name = Path(pdf_path).stem
            excel_path = output_dir / f"{pdf_name}_extraido.xlsx"

            # Extrair dados usando função apropriada
            if tipo == "incra":
                dados = extrair_memorial_incra(Path(pdf_path), api_key)
            else:
                dados = extract_table_from_pdf(pdf_path, api_key)

            # Verificar se dados foram extraídos
            if not dados or 'data' not in dados:
                raise ValueError("Nenhum dado foi extraído do PDF")

            if not dados['data']:
                raise ValueError("PDF extraído, mas tabela de dados está vazia")

            # Criar arquivo Excel
            create_excel_file(dados, str(excel_path))

            # Verificar se o arquivo foi criado
            if not excel_path.exists():
                raise RuntimeError(f"Arquivo Excel não foi criado em: {excel_path}\n"
                                 f"Verifique permissões no diretório: {output_dir}")

            # Verificar se o arquivo tem conteúdo
            file_size = excel_path.stat().st_size
            if file_size == 0:
                raise RuntimeError(f"Arquivo Excel criado mas está vazio: {excel_path}")
            return str(excel_path), dados

        except Exception as e:
            error_msg = f"Erro ao extrair PDF para Excel: {str(e)}"
            print(f"❌ {error_msg}")
            raise RuntimeError(error_msg) from e

    def _ler_dados_excel(self, excel_path: str) -> Dict:
        """
        Lê dados estruturados de um arquivo Excel gerado pela extração.

        Args:
            excel_path: Caminho do arquivo Excel

        Returns:
            Dicionário com estrutura padronizada dos dados
        """
        wb = load_workbook(excel_path)
        ws = wb.active

        dados = {
            "header_row1": ["VÉRTICE", "SEGMENTO VANTE"],
            "header_row2": ["Código", "Longitude", "Latitude", "Altitude (m)",
                           "Código", "Azimute", "Dist. (m)", "Confrontações"],
            "data": []
        }

        # Ler dados a partir da linha 3 (linhas 1 e 2 são cabeçalhos)
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0]:  # Se tem código no vértice
                dados["data"].append(list(row))

        wb.close()
        return dados

    # ========== FIM NOVAS FUNÇÕES V3 ==========

    def _carregar_pdf_como_imagens(self, pdf_path: str, rotacionar_90: bool = False) -> List[Image.Image]:
        """
        Converte um PDF em lista de imagens PIL.
        
        Args:
            pdf_path: Caminho do arquivo PDF
            rotacionar_90: Se True, rotaciona as imagens 90 graus (para INCRA)
            
        Returns:
            Lista de objetos PIL.Image
        """
        try:
            self._atualizar_status(f"Convertendo PDF: {Path(pdf_path).name}...")
            
            # Converter PDF para imagens
            images = convert_from_path(pdf_path, dpi=200)
            
            # Rotacionar se necessário (INCRA em paisagem)
            if rotacionar_90:
                self._atualizar_status(f"Rotacionando imagens do INCRA...")
                images = [img.rotate(-90, expand=True) for img in images]
                
            return images
            
        except Exception as e:
            raise Exception(f"Erro ao processar PDF {Path(pdf_path).name}: {str(e)}")
            
    def _construir_prompt_gemini(self, incluir_projeto: bool = False, incluir_memorial: bool = True) -> List:
        """
        Constrói o prompt multimodal para a API do Gemini.
        
        Args:
            incluir_projeto: Se True, inclui as imagens do projeto na análise
            incluir_memorial: Se True, inclui as imagens do memorial na análise
            
        Returns:
            Lista contendo strings de texto e objetos PIL.Image
        """
        prompt = [
            "Você é um assistente ESPECIALISTA em análise de documentos de georreferenciamento de imóveis rurais para cartórios no Brasil.",
            "\n═══════════════════════════════════════════════════════════",
            "\n=== INSTRUÇÕES CRÍTICAS DE EXTRAÇÃO ===",
            "\n═══════════════════════════════════════════════════════════",
            "\n",
            "\n⚠️⚠️⚠️ ATENÇÃO MÁXIMA - ERROS COMUNS A EVITAR ⚠️⚠️⚠️",
            "\n",
            "\n❌ NÃO CONFUNDA:",
            "\n1. CPF (formato XXX.XXX.XXX-XX) ≠ Código INCRA (formato XXX.XXX.XXX.XXX-X)",
            "\n   • CPF: 765.656.618-04 (pessoa física)",
            "\n   • Código INCRA: 951.742.953-1 (imóvel rural)",
            "\n   • São COMPLETAMENTE diferentes!",
            "\n",
            "\n2. Nomes de proprietários DIFERENTES = STATUS ❌ (não ⚠️!)",
            "\n   • 'PAULO EDUARDO HOTZ' ≠ 'Paulo Gemma Henge'",
            "\n   • São PESSOAS DIFERENTES! Marque como ❌ ERRO GRAVE!",
            "\n   • Não diga 'pequena divergência' - é ERRO TOTAL!",
            "\n",
            "\n3. Memorial em texto corrido TEM perímetro - PROCURE NO TEXTO!",
            "\n   • Busque por: 'perímetro de X metros' ou 'perímetro de X m'",
            "\n   • Exemplo: 'Perímetro (m): 3.873,67 m' ou 'perímetro de 3.873,67 metros'",
            "\n   • Se encontrar, extraia! Não diga 'Não encontrado'!",
            "\n",
            "\n4. Projeto/Planta tem TABELAS - LEIA A TABELA COMPLETA!",
            "\n   • Projetos em PDF digital têm tabelas de coordenadas",
            "\n   • Procure por colunas: Código, Longitude, Latitude, Altitude",
            "\n   • Ou: Código, E (Este), N (Norte)",
            "\n   • EXTRAIA TODOS OS VÉRTICES DA TABELA!",
            "\n   • Não invente coordenadas - copie da tabela!",
            "\n",
            "\n**FORMATO DOS DOCUMENTOS:**",
            "\n1. 📋 INCRA: Dados em TABELAS - extraia TODAS as células com precisão",
            "\n2. 📄 MEMORIAL: Dados em TEXTO CORRIDO - ⚠️ CRÍTICO: LEIA LETRA POR LETRA!",
            "\n   • O Memorial é um texto em PROSA (parágrafos longos)",
            "\n   • As informações estão DISPERSAS e MISTURADAS no texto",
            "\n   • Você DEVE ler com EXTREMA ATENÇÃO cada palavra",
            "\n   • NÃO invente informações - copie EXATAMENTE como está escrito",
            "\n   • Exemplo: Se está 'NCXC-P-1032', escreva EXATAMENTE 'NCXC-P-1032'",
            "\n   • ⚠️ NÃO troque letras! NCXC ≠ NXCX ≠ NCXX ≠ NCCX",
            "\n3. 🗺️ PROJETO/PLANTA: ",
            "\n   • Se for PDF DIGITAL (texto selecionável): TEM TABELAS! Leia-as!",
            "\n   • Se for ESCANEADO (imagem): Extraia visualmente",
            "\n   • Procure por 'Tabela de Coordenadas' ou grade com vértices",
            "\n   • NO PROJETO que você está analisando agora: HÁ UMA TABELA NO CANTO!",
            "\n",
            "\n**⚠️ ATENÇÃO MÁXIMA AO LER MEMORIAL DESCRITIVO:**",
            "\nO Memorial é um TEXTO LONGO onde as informações aparecem assim:",
            "\n'...inicia-se no vértice NCXC-P-1032, de coordenadas (Longitude: -48°40'19,003\", Latitude: -21°00'03,754\"...'",
            "\nOU:",
            "\n'Perímetro (m): 3.873,67 m'",
            "\n",
            "\nVocê DEVE:",
            "\n✅ Ler palavra por palavra, letra por letra",
            "\n✅ Copiar códigos EXATAMENTE: NCXC-P-1032 (não invente NXCX ou similar)",
            "\n✅ Extrair coordenadas completas (Longitude, Latitude, Altitude se houver)",
            "\n✅ Identificar TODOS os vértices mesmo que estejam em parágrafos diferentes",
            "\n✅ Procurar informações em TODO o texto (começo, meio, fim)",
            "\n✅ Buscar 'Perímetro' ou 'perímetro' no texto - NÃO diga 'não encontrado' sem procurar!",
            "\n",
            "\n**⚠️ ATENÇÃO MÁXIMA AO LER PROJETO/PLANTA:**",
            "\n",
            "\n🎯 O PROJETO TEM UMA TABELA! Exemplo:",
            "\n```",
            "\nCódigo      | Longitude        | Latitude         | Altitude",
            "\nAKE-V-0166  | 48°34'14,782\" W | 20°50'45,291\" S | 532,78",
            "\nAKE-M-1028  | 48°34'13,821\" W | 20°50'46,394\" S | 533,92",
            "\n```",
            "\n",
            "\nOU formato UTM:",
            "\n```",
            "\nCódigo      | E (Este)  | N (Norte)",
            "\nAKE-V-0166  | 741319    | 7696237",
            "\n```",
            "\n",
            "\nVocê DEVE:",
            "\n✅ Procurar pela tabela (geralmente no canto ou no topo)",
            "\n✅ Ler TODAS as linhas da tabela",
            "\n✅ Extrair TODOS os vértices listados",
            "\n✅ Copiar coordenadas EXATAMENTE como na tabela",
            "\n✅ Se houver 26 vértices na tabela, liste os 26!",
            "\n✅ NÃO invente coordenadas - só o que está na tabela",
            "\n",
            "\n**EQUIVALÊNCIAS SEMÂNTICAS (MUITO IMPORTANTE!):**",
            "\n- '19,0211 ha' = 'Área: 19.0211 hectares' = 'ÁREA TOTAL (ha): 19,0211'",
            "\n- 'José da Silva' = 'Sr. José da Silva' = 'JOSÉ DA SILVA' = 'Jose da Silva'",
            "\n- Vírgula e ponto decimal são equivalentes: 19,02 = 19.02",
            "\n- Espaços e formatação diferentes não importam",
            "\n",
            "\n**⚠️ MAS ATENÇÃO - QUANDO NÃO É EQUIVALENTE:**",
            "\n- 'PAULO EDUARDO HOTZ' ≠ 'Paulo Gemma Henge' → São PESSOAS DIFERENTES! Status = ❌",
            "\n- '951.742.953-1' ≠ '765.656.618-04' → Um é Código INCRA, outro é CPF! Status = ❌",
            "\n- '3.873,67 m' ≠ 'Não encontrado' → Um tem valor, outro não! Status = ❌",
            "\n- 'Latitude/Longitude' ≠ 'UTM' → Sistemas DIFERENTES! Status = ⚠️",
            "\n",
            "\n**⚠️ ATENÇÃO ESPECIAL - INFORMAÇÕES PARCIAIS:**",
            "\n- Se um documento tem TEXTO PARCIAL de outro, isso NÃO é igual!",
            "\n- Exemplo ERRADO de considerar igual:",
            "\n  • INCRA: 'Estrada Municipal'",
            "\n  • Memorial: 'Estrada Municipal que liga o distrito de São José ao centro'",
            "\n  → Isso é DIFERENTE! O Memorial tem informação ADICIONAL importante!",
            "\n- Quando encontrar casos assim, marque como <span class='status-alerta'>⚠️</span>",
            "\n- E adicione observação: 'VERIFICAR: Um documento tem informação mais completa'",
            "\n- O usuário DEVE verificar manualmente se a informação adicional é relevante",
            "\n",
            "\n**DADOS QUE VOCÊ DEVE EXTRAIR DE CADA DOCUMENTO:**",
            "\n",
            "\n✅ **DADOS BÁSICOS:**",
            "\n   • Proprietário(s) - nome completo EXATO",
            "\n   • Nome do Imóvel/Propriedade",
            "\n   • Matrícula(s) do cartório",
            "\n   • Município e Estado (UF)",
            "\n   • Código INCRA (código de certificação) - NÃO CONFUNDA COM CPF!",
            "\n   • CCIR (se houver)",
            "\n   • Cartório/CNS",
            "\n",
            "\n✅ **DADOS TÉCNICOS:**",
            "\n   • Área Total em hectares (todas as casas decimais)",
            "\n   • Perímetro em metros - BUSQUE NO TEXTO DO MEMORIAL!",
            "\n   • Sistema de coordenadas (UTM/Geográfico/SIRGAS)",
            "\n   • Datum (SIRGAS2000, SAD69, etc)",
            "\n",
            "\n✅ **VÉRTICES E COORDENADAS - ⚠️ MÁXIMA ATENÇÃO:**",
            "\n   • TODOS os vértices (V1, V2, V3, V4, V5, V6...)",
            "\n   • Códigos COMPLETOS dos vértices (ex: NCXC-P-1032, YGGA-M-0046, AKE-V-0166)",
            "\n   • ⚠️ COPIE O CÓDIGO EXATAMENTE LETRA POR LETRA!",
            "\n   • Coordenadas COMPLETAS de cada vértice:",
            "\n     - Longitude (ex: -48°40'19,003\") OU E=741319 (UTM)",
            "\n     - Latitude (ex: -21°00'03,754\") OU N=7696237 (UTM)",
            "\n     - Altitude se houver (ex: 509,05 m)",
            "\n   • CRÍTICO: Não omita vértices! Liste TODOS que encontrar!",
            "\n   • No Memorial, os vértices aparecem assim:",
            "\n     'vértice NCXC-P-1032, de coordenadas (Longitude: -48°40'19,003\", Latitude: -21°00'03,754\"...'",
            "\n     ou",
            "\n     '12,68 m até o vértice NCXC-P-1033, de coordenadas...'",
            "\n   • No Projeto, os vértices estão em TABELAS:",
            "\n     Procure por tabela com colunas: Código | Longitude | Latitude | Altitude",
            "\n     Ou: Código | E | N",
            "\n",
            "\n✅ **CONFRONTANTES/LIMITES:**",
            "\n   • Norte: [quem/o quê]",
            "\n   • Sul: [quem/o quê]",
            "\n   • Leste: [quem/o quê]",
            "\n   • Oeste: [quem/o quê]",
            "\n",
            "\n--- INÍCIO DOCUMENTO INCRA ---",
            "\n",
            "\n🚨🚨🚨 ALERTA CRÍTICO - CÓDIGOS DOS VÉRTICES 🚨🚨🚨",
            "\n",
            "\n⚠️⚠️⚠️ PROBLEMA COMUM DE OCR:",
            "\nO OCR frequentemente CONFUNDE a letra 'K' com 'M'!",
            "\n",
            "\n❌ ERRO GRAVÍSSIMO:",
            "\n   AME-V-0166  ← ERRADO! (K virou M)",
            "\n   AME-M-1028  ← ERRADO! (K virou M)",
            "\n   AME-P-3567  ← ERRADO! (K virou M)",
            "\n",
            "\n✅ CÓDIGOS CORRETOS:",
            "\n   AKE-V-0166  ← CORRETO! (com K)",
            "\n   AKE-M-1028  ← CORRETO! (com K)",
            "\n   AKE-P-3567  ← CORRETO! (com K)",
            "\n",
            "\n🔍 COMO IDENTIFICAR:",
            "\nOlhe com ATENÇÃO EXTREMA para as primeiras 3 letras do código:",
            "\n• Se parece 'AME' → É ERRO! Deve ser 'AKE'",
            "\n• Se parece 'AXE' → É ERRO! Deve ser 'AKE'",
            "\n• Se parece 'AKF' → É ERRO! Deve ser 'AKE'",
            "\n",
            "\n💡 DICA:",
            "\nNeste documento, o código de credenciamento é 'AKE'.",
            "\nPORTANTO, TODOS os vértices começam com 'AKE-'!",
            "\n",
            "\n⚠️ NUNCA NUNCA NUNCA escreva 'AME'!",
            "\n⚠️ SEMPRE escreva 'AKE' com a letra K!",
            "\n",
            "\n🎯 EXTRAÇÃO ESPECÍFICA DO INCRA - INSTRUÇÕES CIRÚRGICAS",
            "\n",
            "\n════════════════════════════════════════════════════════════",
            "\n                PARTE 1: DADOS CADASTRAIS                   ",
            "\n════════════════════════════════════════════════════════════",
            "\n",
            "\nExtraia APENAS as seguintes informações, NESTA ORDEM:",
            "\n",
            "\n1️⃣ **Denominação:**",
            "\n   • PROCURE: Linha que começa com 'Denominação:'",
            "\n   • EXTRAIA: SOMENTE o nome do imóvel",
            "\n   • REMOVA: Qualquer menção a 'Área X', 'Matrícula', números",
            "\n   • EXEMPLO:",
            "\n     ❌ Errado: 'Fazenda Monte Rosa - Área 2 – Matrícula n° 27.935'",
            "\n     ✅ Correto: 'Fazenda Monte Rosa'",
            "\n",
            "\n2️⃣ **Proprietário(a):**",
            "\n   • PROCURE: Linha que começa com 'Proprietário(a):'",
            "\n   • EXTRAIA: Nome completo do proprietário",
            "\n   • EXEMPLO: 'RENÊ EDUARDO HOTZ'",
            "\n",
            "\n3️⃣ **Matrícula do imóvel:**",
            "\n   • PROCURE: Linha 'Matrícula do imóvel:'",
            "\n   • ATENÇÃO: Pode ter continuação na página 3!",
            "\n   • EXTRAIA: TODOS os números de matrícula",
            "\n   • EXEMPLO: '28625, 28626, 27935, 27936, 11798'",
            "\n   • LEMBRE: Procurar também: 'continuação da página 1: ...'",
            "\n",
            "\n4️⃣ **Município/UF:**",
            "\n   • PROCURE: 'Município/UF:'",
            "\n   • EXTRAIA: Nome do município e UF",
            "\n   • EXEMPLO: 'Bebedouro-SP'",
            "\n",
            "\n5️⃣ **Código de credenciamento:**",
            "\n   • PROCURE: 'Código de credenciamento:'",
            "\n   • EXTRAIA: O código (geralmente 3 letras)",
            "\n   • EXEMPLO: 'AKE'",
            "\n",
            "\n6️⃣ **Código INCRA/SNCR:**",
            "\n   • PROCURE: 'Código INCRA/SNCR:'",
            "\n   • EXTRAIA: Código completo",
            "\n   • EXEMPLO: '6120730013504'",
            "\n   • ⚠️ NÃO confunda com CPF!",
            "\n",
            "\n7️⃣ **Área (Sistema Geodésico Local):**",
            "\n   • PROCURE: 'Área (Sistema Geodésico Local):'",
            "\n   • EXTRAIA: Valor e unidade",
            "\n   • EXEMPLO: '68,7187 ha'",
            "\n",
            "\n8️⃣ **Perímetro (m):**",
            "\n   • PROCURE: 'Perímetro (m):'",
            "\n   • EXTRAIA: Valor em metros",
            "\n   • EXEMPLO: '3.873,67 m'",
            "\n",
            "\n════════════════════════════════════════════════════════════",
            "\n              PARTE 2: TABELA DE COORDENADAS                ",
            "\n════════════════════════════════════════════════════════════",
            "\n",
            "\n📊 LOCALIZAÇÃO DA TABELA:",
            "\n   • Título: 'DESCRIÇÃO DA PARCELA'",
            "\n   • Tem 2 seções lado a lado:",
            "\n     - VÉRTICE (esquerda): Código, Longitude, Latitude, Altitude",
            "\n     - SEGMENTO VANTE (direita): Código, Azimute, Dist.(m), Confrontações",
            "\n",
            "\n⚠️ INSTRUÇÕES CRÍTICAS PARA LER A TABELA:",
            "\n",
            "\n1. LOCALIZE a tabela 'DESCRIÇÃO DA PARCELA'",
            "\n",
            "\n2. A tabela tem este formato:",
            "\n┌─────────────┬────────────────┬────────────────┬─────────────┐",
            "\n│ VÉRTICE                                                      │",
            "\n├─────────────┼────────────────┼────────────────┼─────────────┤",
            "\n│ Código      │ Longitude      │ Latitude       │ Altitude(m) │",
            "\n├─────────────┼────────────────┼────────────────┼─────────────┤",
            "\n│ AKE-V-0166  │ -48°34'14,782\" │ -20°50'45,291\" │ 532,78      │",
            "\n└─────────────┴────────────────┴────────────────┴─────────────┘",
            "\n",
            "\n┌─────────────┬─────────┬──────────┬─────────────────────────┐",
            "\n│ SEGMENTO VANTE                                              │",
            "\n├─────────────┼─────────┼──────────┼─────────────────────────┤",
            "\n│ Código      │ Azimute │ Dist.(m) │ Confrontações           │",
            "\n├─────────────┼─────────┼──────────┼─────────────────────────┤",
            "\n│ AKE-M-1028  │ 140°40' │ 43,85    │ CNS: 12.102-0 | Mat...  │",
            "\n└─────────────┴─────────┴──────────┴─────────────────────────┘",
            "\n",
            "\n3. COPIE os códigos dos vértices EXATAMENTE:",
            "\n   • Exemplo: AKE-V-0166, AKE-M-1028, AKE-P-3567",
            "\n   • ⚠️ NÃO troque letras: AKE ≠ AME ≠ AXE ≠ AKF",
            "\n   • ⚠️ NÃO troque números: 1028 ≠ 1008 ≠ 1128",
            "\n   • ⚠️ Mantenha hífens e letras: AKE-P-3567 (não AKE P 3567)",
            "\n",
            "\n4. COPIE as coordenadas COM TODOS OS SÍMBOLOS:",
            "\n   • Longitude: -48°34'14,782\" (sinal, °, ', \")",
            "\n   • Latitude: -20°50'45,291\" (sinal, °, ', \")",
            "\n   • Altitude: 532,78 (número com vírgula)",
            "\n   • Azimute: 140°40' (graus e minutos)",
            "\n   • Distância: 43,85 (número com vírgula)",
            "\n",
            "\n5. REPRODUZA A TABELA COMPLETA:",
            "\n   • ⚠️ A tabela continua em MÚLTIPLAS PÁGINAS!",
            "\n   • Página 1: Primeiros ~16 vértices",
            "\n   • Página 2: Vértices restantes (~10)",
            "\n   • TOTAL: ~26 vértices",
            "\n   • COPIE TODOS! Não pare na página 1!",
            "\n",
            "\n6. MANTENHA A FORMATAÇÃO:",
            "\n   • Use espaços/tabs para alinhar colunas",
            "\n   • Separe seções (VÉRTICE e SEGMENTO VANTE)",
            "\n   • Mantenha símbolos especiais (°, ', \")",
            "\n",
            "\n7. CONFRONTANTES DO INCRA:",
            "\n   • Os confrontantes estão na coluna 'Confrontações' da tabela",
            "\n   • Exemplo: 'CNS: 12.102-0 | Mat. 28309'",
            "\n   • Exemplo: 'Estrada Municipal - BBD 315'",
            "\n   • Exemplo: 'CNS: 12.102-0 | Mat. 34685 | Córrego Lambari'",
            "\n   • ⚠️ NÃO extraia nomes de pessoas!",
            "\n   • ✅ Extraia: Matrícula, nome da estrada, córrego, etc.",
            "\n",
            "\n════════════════════════════════════════════════════════════",
            "\n                    FORMATO DE SAÍDA                         ",
            "\n════════════════════════════════════════════════════════════",
            "\n",
            "\nApresente no seguinte formato:",
            "\n",
            "\n**DADOS CADASTRAIS:**",
            "\nDenominação: [valor]",
            "\nProprietário(a): [valor]",
            "\nMatrícula do imóvel: [valor]",
            "\nMunicípio/UF: [valor]",
            "\nCódigo de credenciamento: [valor]",
            "\nCódigo INCRA/SNCR: [valor]",
            "\nÁrea (Sistema Geodésico Local): [valor]",
            "\nPerímetro (m): [valor]",
            "\n",
            "\n**TABELA DE COORDENADAS:**",
            "\n[Reproduza a tabela completa aqui, mantendo formatação]",
            "\n",
            "\nExtraia CADA dado de CADA célula com MÁXIMA PRECISÃO!",
        ]
        
        # Adicionar imagens do INCRA
        prompt.extend(self.incra_images)
        prompt.append("\n--- FIM DOCUMENTO INCRA ---")
        
        # Adicionar imagens do Memorial se necessário
        if incluir_memorial and self.memorial_images:
            prompt.append("\n--- INÍCIO MEMORIAL DESCRITIVO ---")
            prompt.append("\n⚠️ ATENÇÃO: Este documento tem TEXTO CORRIDO.")
            prompt.append("\nLeia TODO o conteúdo com cuidado.")
            prompt.append("\nAs informações estão espalhadas em parágrafos diferentes.")
            prompt.extend(self.memorial_images)
            prompt.append("\n--- FIM MEMORIAL DESCRITIVO ---")
        
        # Adicionar imagens do Projeto se solicitado
        if incluir_projeto and self.projeto_images:
            prompt.append("\n--- INÍCIO PROJETO/PLANTA ---")
            prompt.append("\n🎯 ATENÇÃO ESPECIAL PARA ESTE PROJETO:")
            prompt.append("\nEste é um PDF DIGITAL (não escaneado) - ele contém TABELAS DE DADOS!")
            prompt.append("\n")
            prompt.append("\n📊 ONDE ESTÁ A TABELA:")
            prompt.append("\nProcure por uma tabela com o título:")
            prompt.append("\n'Tabela de Coordenadas - Altitudes - Azimutes - Distâncias'")
            prompt.append("\n")
            prompt.append("\nA tabela tem as seguintes colunas:")
            prompt.append("\n┌──────────┬────────────────┬────────────────┬────────────┐")
            prompt.append("\n│ Código   │ Longitude      │ Latitude       │ Altitude   │")
            prompt.append("\n├──────────┼────────────────┼────────────────┼────────────┤")
            prompt.append("\n│ AKE-V... │ 48°34'14,782\" W│ 20°50'45,291\" S│ 532,78     │")
            prompt.append("\n└──────────┴────────────────┴────────────────┴────────────┘")
            prompt.append("\n")
            prompt.append("\n⚠️ INSTRUÇÕES CRÍTICAS DE EXTRAÇÃO:")
            prompt.append("\n")
            prompt.append("\n1. 🔍 LOCALIZE a tabela completa")
            prompt.append("\n   • Geralmente está no CANTO ESQUERDO da página")
            prompt.append("\n   • Ou na parte SUPERIOR")
            prompt.append("\n   • Título: 'Tabela de Coordenadas...'")
            prompt.append("\n")
            prompt.append("\n2. 📖 LEIA LINHA POR LINHA")
            prompt.append("\n   • Primeira linha: Cabeçalhos (Código, Longitude, Latitude, Altitude)")
            prompt.append("\n   • Depois: TODAS as linhas de dados")
            prompt.append("\n   • Pode ter 20, 26, 30 ou mais vértices!")
            prompt.append("\n")
            prompt.append("\n3. ✍️ COPIE EXATAMENTE")
            prompt.append("\n   • Código do vértice: AKE-V-0166, AKE-M-1028, AKE-P-3567...")
            prompt.append("\n   • Longitude: 48°34'14,782\" W (com graus, minutos, segundos E direção)")
            prompt.append("\n   • Latitude: 20°50'45,291\" S (com graus, minutos, segundos E direção)")
            prompt.append("\n   • Altitude: 532,78 (número simples)")
            prompt.append("\n")
            prompt.append("\n4. ⚠️ NÃO CONFUNDA:")
            prompt.append("\n   • ❌ NÃO pegue números do DESENHO (ex: E=741319 N=7696237)")
            prompt.append("\n   • ❌ NÃO pegue números das LEGENDAS")
            prompt.append("\n   • ❌ NÃO pegue números dos CARIMBOS")
            prompt.append("\n   • ✅ SÓ pegue da TABELA DE COORDENADAS!")
            prompt.append("\n")
            prompt.append("\n5. 📝 LISTE TODOS")
            prompt.append("\n   • Se a tabela tem 26 vértices, liste os 26!")
            prompt.append("\n   • Não omita nenhum vértice")
            prompt.append("\n   • Não pare em 3-4 vértices")
            prompt.append("\n")
            prompt.append("\n💡 EXEMPLO CORRETO DE EXTRAÇÃO:")
            prompt.append("\nVértice AKE-V-0166:")
            prompt.append("\n  • Longitude: 48°34'14,782\" W")
            prompt.append("\n  • Latitude: 20°50'45,291\" S")
            prompt.append("\n  • Altitude: 532,78 m")
            prompt.append("\n")
            prompt.append("\nVértice AKE-M-1028:")
            prompt.append("\n  • Longitude: 48°34'13,821\" W")
            prompt.append("\n  • Latitude: 20°50'46,394\" S")
            prompt.append("\n  • Altitude: 533,92 m")
            prompt.append("\n")
            prompt.append("\n... (continua para TODOS os vértices da tabela)")
            prompt.append("\n")
            prompt.append("\n❌ EXEMPLO ERRADO (NÃO FAÇA ISSO):")
            prompt.append("\n'E=741319 N=7696237' ← Isso é do DESENHO, não da tabela!")
            prompt.append("\n")
            prompt.extend(self.projeto_images)
            prompt.append("\n--- FIM PROJETO/PLANTA ---")
            
        # Instruções de formato de saída - HTML PROFISSIONAL COM CORES
        
        # Determinar quais documentos foram fornecidos
        docs_fornecidos = []
        if self.incra_images:
            docs_fornecidos.append("INCRA")
        if incluir_memorial and self.memorial_images:
            docs_fornecidos.append("MEMORIAL")
        if incluir_projeto and self.projeto_images:
            docs_fornecidos.append("PROJETO")
        
        docs_texto = " + ".join(docs_fornecidos)
        
        instrucoes_saida = (
            "\n\n"
            "\n════════════════════════════════════════════════════════════════════"
            "\n                    FORMATO DO RELATÓRIO HTML                       "
            "\n════════════════════════════════════════════════════════════════════"
            "\n"
            f"\n🎯 DOCUMENTOS SENDO COMPARADOS: {docs_texto}"
            "\n"
            "\n⚠️⚠️⚠️ REGRA CRÍTICA DE FORMATAÇÃO:"
            "\n"
            "\n1️⃣ SOMENTE inclua no relatório os documentos que foram fornecidos!"
            "\n"
        )
        
        # Adicionar instruções específicas baseadas nos documentos
        if incluir_memorial and not incluir_projeto:
            instrucoes_saida += (
                "\n   Você está comparando: INCRA + MEMORIAL"
                "\n   • Tabela deve ter 3 colunas: DADO | INCRA | MEMORIAL | STATUS"
                "\n   • NÃO mencione 'Projeto' ou 'Planta' em lugar nenhum"
                "\n   • NÃO crie coluna 'PROJETO'"
                "\n"
            )
        elif incluir_projeto and not incluir_memorial:
            instrucoes_saida += (
                "\n   Você está comparando: INCRA + PROJETO"
                "\n   • Tabela deve ter 3 colunas: DADO | INCRA | PROJETO | STATUS"
                "\n   • NÃO mencione 'Memorial' ou 'Memorial Descritivo' em lugar nenhum"
                "\n   • NÃO crie coluna 'MEMORIAL'"
                "\n"
            )
        else:  # Todos os 3
            instrucoes_saida += (
                "\n   Você está comparando: INCRA + MEMORIAL + PROJETO"
                "\n   • Tabela deve ter 4 colunas: DADO | INCRA | MEMORIAL | PROJETO | STATUS"
                "\n"
            )
        
        instrucoes_saida += (
            "\n2️⃣ Para documentos NÃO fornecidos:"
            "\n   • NÃO crie coluna para eles"
            "\n   • NÃO escreva 'N/A' ou 'Não fornecido'"
            "\n   • SIMPLESMENTE omita essa coluna"
            "\n"
            "\n3️⃣ Estrutura da tabela:"
        )
        
        # Cabeçalho da tabela baseado nos documentos
        if incluir_memorial and not incluir_projeto:
            instrucoes_saida += (
                "\n   <thead><tr>"
                "\n       <th>DADO</th>"
                "\n       <th>INCRA</th>"
                "\n       <th>MEMORIAL</th>"
                "\n       <th>STATUS</th>"
                "\n   </tr></thead>"
            )
        elif incluir_projeto and not incluir_memorial:
            instrucoes_saida += (
                "\n   <thead><tr>"
                "\n       <th>DADO</th>"
                "\n       <th>INCRA</th>"
                "\n       <th>PROJETO</th>"
                "\n       <th>STATUS</th>"
                "\n   </tr></thead>"
            )
        else:
            instrucoes_saida += (
                "\n   <thead><tr>"
                "\n       <th>DADO</th>"
                "\n       <th>INCRA</th>"
                "\n       <th>MEMORIAL</th>"
                "\n       <th>PROJETO</th>"
                "\n       <th>STATUS</th>"
                "\n   </tr></thead>"
            )
        
        instrucoes_saida += (
            "\n"
            "\n⚠️ IMPORTANTE: Gere um relatório em HTML completo e profissional."
            "\nUse CSS inline para cores, estilos e organização visual perfeita."
            "\nCada seção deve ter cores diferentes para fácil identificação."
            "\n"
            "\nGere EXATAMENTE este formato HTML (adapte os dados):"
            "\n"
            "\n```html"
            "\n<!DOCTYPE html>"
            "\n<html lang='pt-BR'>"
            "\n<head>"
            "\n    <meta charset='UTF-8'>"
            "\n    <meta name='viewport' content='width=device-width, initial-scale=1.0'>"
            "\n    <title>Relatório de Consistência - Georreferenciamento</title>"
            "\n    <style>"
            "\n        body { font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }"
            "\n        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 30px; box-shadow: 0 0 10px rgba(0,0,0,0.1); }"
            "\n        h1 { color: #2c3e50; border-bottom: 4px solid #3498db; padding-bottom: 10px; }"
            "\n        h2 { color: #34495e; margin-top: 30px; padding: 10px; border-left: 5px solid #3498db; background: #ecf0f1; }"
            "\n        .resumo { background: #e8f5e9; padding: 20px; border-left: 5px solid #4caf50; margin: 20px 0; font-size: 16px; }"
            "\n        .resumo.alerta { background: #fff3e0; border-left-color: #ff9800; }"
            "\n        .resumo.erro { background: #ffebee; border-left-color: #f44336; }"
            "\n        table { width: 100%; border-collapse: collapse; margin: 20px 0; font-size: 14px; }"
            "\n        th { background: #3498db; color: white; padding: 12px; text-align: left; font-weight: bold; }"
            "\n        td { padding: 10px; border: 1px solid #ddd; }"
            "\n        tr:nth-child(even) { background: #f9f9f9; }"
            "\n        tr:hover { background: #f0f0f0; }"
            "\n        .status-ok { color: #4caf50; font-weight: bold; font-size: 18px; }"
            "\n        .status-alerta { color: #ff9800; font-weight: bold; font-size: 18px; }"
            "\n        .status-erro { color: #f44336; font-weight: bold; font-size: 18px; }"
            "\n        .secao-cadastro th { background: #2196f3; }"
            "\n        .secao-tecnico th { background: #009688; }"
            "\n        .secao-vertices th { background: #673ab7; }"
            "\n        .secao-confrontantes th { background: #ff5722; }"
            "\n        .secao-erros { background: #ffebee; padding: 15px; border-left: 5px solid #f44336; margin: 20px 0; }"
            "\n        .secao-alertas { background: #fff3e0; padding: 15px; border-left: 5px solid #ff9800; margin: 20px 0; }"
            "\n        .secao-ok { background: #e8f5e9; padding: 15px; border-left: 5px solid #4caf50; margin: 20px 0; }"
            "\n        .parecer { padding: 20px; margin: 20px 0; border: 3px solid; font-size: 16px; font-weight: bold; }"
            "\n        .parecer-aprovado { background: #e8f5e9; border-color: #4caf50; color: #2e7d32; }"
            "\n        .parecer-ressalvas { background: #fff3e0; border-color: #ff9800; color: #e65100; }"
            "\n        .parecer-reprovado { background: #ffebee; border-color: #f44336; color: #c62828; }"
            "\n        .legenda { background: #ecf0f1; padding: 15px; margin: 20px 0; border-radius: 5px; }"
            "\n        .analise { font-style: italic; color: #555; margin: 10px 0; padding: 10px; background: #f9f9f9; }"
            "\n    </style>"
            "\n</head>"
            "\n<body>"
            "\n<div class='container'>"
            "\n"
            "\n<h1>📊 RELATÓRIO DE CONSISTÊNCIA - GEORREFERENCIAMENTO</h1>"
            "\n"
            "\n<!-- RESUMO EXECUTIVO -->"
            "\n<h2>🎯 RESUMO EXECUTIVO</h2>"
            "\n<div class='resumo'> <!-- Use classe 'alerta' ou 'erro' se houver problemas -->"
            "\n[Em 2-3 frases diretas: os documentos estão consistentes ou há erros?]"
            "\n</div>"
            "\n"
            "\n<!-- SEÇÃO 1: DADOS CADASTRAIS -->"
            "\n<h2>📋 1. DADOS CADASTRAIS</h2>"
            "\n<table class='secao-cadastro'>"
            "\n<thead>"
            "\n    <tr>"
            "\n        <th>DADO</th>"
            "\n        [COLUNAS DOS DOCUMENTOS FORNECIDOS]"
            "\n        <th style='text-align:center;'>STATUS</th>"
            "\n    </tr>"
            "\n</thead>"
            "\n<tbody>"
            "\n    <tr>"
            "\n        <td><strong>Proprietário(s)</strong></td>"
            "\n        [DADOS DE CADA DOCUMENTO]"
            "\n        <td style='text-align:center;'><span class='status-ok'>✅</span></td>"
            "\n    </tr>"
            "\n    <!-- Repetir para: Nome do Imóvel, Matrícula(s), Município, UF, Código INCRA, etc -->"
            "\n    <tr>"
            "\n        <td><strong>UF</strong></td>"
            "\n        <td>[extrair]</td>"
            "\n        <td>[extrair]</td>"
            "\n        <td>[extrair/N/A]</td>"
            "\n        <td style='text-align:center;'><span class='status-ok'>✅</span></td>"
            "\n    </tr>"
            "\n    <tr>"
            "\n        <td><strong>Código INCRA</strong></td>"
            "\n        <td>[extrair]</td>"
            "\n        <td>[extrair]</td>"
            "\n        <td>[extrair/N/A]</td>"
            "\n        <td style='text-align:center;'><span class='status-ok'>✅</span></td>"
            "\n    </tr>"
            "\n    <tr>"
            "\n        <td><strong>CCIR</strong></td>"
            "\n        <td>[extrair]</td>"
            "\n        <td>[extrair]</td>"
            "\n        <td>[extrair/N/A]</td>"
            "\n        <td style='text-align:center;'><span class='status-ok'>✅</span></td>"
            "\n    </tr>"
            "\n</tbody>"
            "\n</table>"
            "\n<p class='analise'><strong>Análise:</strong> [Breve comentário sobre consistência destes dados]</p>"
            "\n"
            "\n<!-- SEÇÃO 2: DADOS TÉCNICOS -->"
            "\n<h2>📐 2. DADOS TÉCNICOS/MENSURAÇÕES</h2>"
            "\n<table class='secao-tecnico'>"
            "\n<thead>"
            "\n    <tr>"
            "\n        <th>DADO</th>"
            "\n        <th>INCRA</th>"
            "\n        <th>MEMORIAL</th>"
            "\n        <th>PROJETO</th>"
            "\n        <th style='text-align:center;'>STATUS</th>"
            "\n    </tr>"
            "\n</thead>"
            "\n<tbody>"
            "\n    <tr>"
            "\n        <td><strong>Área Total (ha)</strong></td>"
            "\n        <td>[X,XXXX]</td>"
            "\n        <td>[X,XXXX]</td>"
            "\n        <td>[X,XXXX/N/A]</td>"
            "\n        <td style='text-align:center;'><span class='status-ok'>✅</span></td>"
            "\n    </tr>"
            "\n    <tr>"
            "\n        <td><strong>Perímetro (m)</strong></td>"
            "\n        <td>[X.XXX,XX]</td>"
            "\n        <td>[X.XXX,XX]</td>"
            "\n        <td>[X.XXX,XX/N/A]</td>"
            "\n        <td style='text-align:center;'><span class='status-ok'>✅</span></td>"
            "\n    </tr>"
            "\n    <tr>"
            "\n        <td><strong>Sistema Coordenadas</strong></td>"
            "\n        <td>[UTM/GEO]</td>"
            "\n        <td>[UTM/GEO]</td>"
            "\n        <td>[UTM/GEO/N/A]</td>"
            "\n        <td style='text-align:center;'><span class='status-ok'>✅</span></td>"
            "\n    </tr>"
            "\n    <tr>"
            "\n        <td><strong>Datum</strong></td>"
            "\n        <td>[SIRGAS]</td>"
            "\n        <td>[SIRGAS]</td>"
            "\n        <td>[SIRGAS/N/A]</td>"
            "\n        <td style='text-align:center;'><span class='status-ok'>✅</span></td>"
            "\n    </tr>"
            "\n    <tr>"
            "\n        <td><strong>Fuso</strong></td>"
            "\n        <td>[22/23]</td>"
            "\n        <td>[22/23]</td>"
            "\n        <td>[22/23/N/A]</td>"
            "\n        <td style='text-align:center;'><span class='status-ok'>✅</span></td>"
            "\n    </tr>"
            "\n</tbody>"
            "\n</table>"
            "\n<p class='analise'><strong>Análise:</strong> [Breve comentário sobre consistência destes dados]</p>"
            "\n"
            "\n<!-- SEÇÃO 3: VÉRTICES -->"
            "\n<h2>🗺️ 3. COORDENADAS DOS VÉRTICES</h2>"
            "\n<p><strong>⚠️ CRÍTICO: Liste TODOS os vértices encontrados!</strong></p>"
            "\n<p><strong>⚠️ COPIE os códigos EXATAMENTE como aparecem no documento!</strong></p>"
            "\n<p style='background:#fff3e0; padding:10px; border-left:3px solid #ff9800;'>"
            "\n<strong>Exemplo de extração do Memorial:</strong><br>"
            "\nSe o texto diz: 'vértice NCXC-P-1032, de coordenadas (Longitude: -48°40'19,003\", Latitude: -21°00'03,754\" e Altitude: 509,05 m)'<br>"
            "\nVocê deve extrair:<br>"
            "\n• Código: <strong>NCXC-P-1032</strong> (exatamente assim!)<br>"
            "\n• Longitude: -48°40'19,003\"<br>"
            "\n• Latitude: -21°00'03,754\"<br>"
            "\n• Altitude: 509,05 m"
            "\n</p>"
            "\n<table class='secao-vertices'>"
            "\n<thead>"
            "\n    <tr>"
            "\n        <th>VÉRTICE</th>"
            "\n        <th>INCRA (Coordenadas)</th>"
            "\n        <th>MEMORIAL (Coordenadas)</th>"
            "\n        <th>PROJETO (Coordenadas)</th>"
            "\n        <th style='text-align:center;'>STATUS</th>"
            "\n    </tr>"
            "\n</thead>"
            "\n<tbody>"
            "\n    <tr>"
            "\n        <td><strong>V1</strong></td>"
            "\n        <td>[E=XXX N=YYY]</td>"
            "\n        <td>[E=XXX N=YYY]</td>"
            "\n        <td>[E=XXX N=YYY/N/A]</td>"
            "\n        <td style='text-align:center;'><span class='status-ok'>✅</span></td>"
            "\n    </tr>"
            "\n    <!-- ADICIONE UMA LINHA PARA CADA VÉRTICE (V2, V3, V4... até o último!) -->"
            "\n    <!-- NÃO OMITA NENHUM VÉRTICE! -->"
            "\n</tbody>"
            "\n</table>"
            "\n<p class='analise'><strong>Análise:</strong> [Comentário sobre consistência das coordenadas]</p>"
            "\n"
            "\n<!-- SEÇÃO 4: CONFRONTANTES -->"
            "\n<h2>🧭 4. CONFRONTANTES/LIMITES</h2>"
            "\n"
            "\n⚠️ INSTRUÇÕES ESPECIAIS PARA CONFRONTANTES:"
            "\n"
            "\n📋 INCRA:"
            "\n   • Os confrontantes do INCRA estão na coluna 'Confrontações' da tabela"
            "\n   • Exemplos:"
            "\n     - 'CNS: 12.102-0 | Mat. 28309'"
            "\n     - 'Estrada Municipal - BBD 315'"
            "\n     - 'CNS: 12.102-0 | Mat. 34685 | Córrego Lambari'"
            "\n   • ⚠️ NÃO extraia nomes de pessoas!"
            "\n   • ✅ Extraia: Matrículas, estradas, córregos, limites"
            "\n   • Liste os confrontantes únicos (sem repetir)"
            "\n"
            "\n📄 MEMORIAL:"
            "\n   • Procure por 'confrontando com' ou 'divisa com'"
            "\n   • Pode estar no texto corrido"
            "\n"
            "\n🗺️ PROJETO:"
            "\n   • Pode estar em legendas ou carimbos"
            "\n   • Ou em texto descritivo"
            "\n"
            "\n<table class='secao-confrontantes'>"
            "\n<thead>"
            "\n    <tr>"
            "\n        <th>DIREÇÃO</th>"
            "\n        [COLUNAS DOS DOCUMENTOS FORNECIDOS]"
            "\n        <th style='text-align:center;'>STATUS</th>"
            "\n    </tr>"
            "\n</thead>"
            "\n<tbody>"
            "\n    <!-- Liste os confrontantes encontrados -->"
            "\n    <!-- Pode não ter direção específica, liste todos encontrados -->"
            "\n</tbody>"
            "\n</table>"
            "\n<p class='analise'><strong>Análise:</strong> [Comentário sobre consistência dos confrontantes]</p>"
            "\n"
            "\n<!-- SEÇÃO 5: DISCREPÂNCIAS CRÍTICAS -->"
            "\n<h2>🚨 5. DISCREPÂNCIAS CRÍTICAS</h2>"
            "\n<div class='secao-erros'>"
            "\n[Se NÃO houver erros graves, escreva:]"
            "\n<p><strong>✅ Nenhuma discrepância crítica identificada.</strong></p>"
            "\n"
            "\n[Se HOUVER erros graves, use esta tabela:]"
            "\n<table>"
            "\n<thead>"
            "\n    <tr style='background:#f44336;'>"
            "\n        <th>TIPO</th><th>CAMPO</th><th>INCRA</th><th>MEMORIAL</th><th>PROJETO</th><th>AÇÃO NECESSÁRIA</th>"
            "\n    </tr>"
            "\n</thead>"
            "\n<tbody>"
            "\n    <tr>"
            "\n        <td><span class='status-erro'>❌</span></td>"
            "\n        <td>[campo]</td>"
            "\n        <td>[valor]</td>"
            "\n        <td>[valor]</td>"
            "\n        <td>[valor]</td>"
            "\n        <td>[o que corrigir]</td>"
            "\n    </tr>"
            "\n</tbody>"
            "\n</table>"
            "\n</div>"
            "\n"
            "\n<!-- SEÇÃO 6: PEQUENAS DIVERGÊNCIAS -->"
            "\n<h2>⚠️ 6. PEQUENAS DIVERGÊNCIAS</h2>"
            "\n<div class='secao-alertas'>"
            "\n[Se NÃO houver diferenças pequenas, escreva:]"
            "\n<p><strong>✅ Nenhuma divergência menor identificada.</strong></p>"
            "\n"
            "\n[Se HOUVER pequenas diferenças, use esta tabela:]"
            "\n<table>"
            "\n<thead>"
            "\n    <tr style='background:#ff9800;'>"
            "\n        <th>TIPO</th><th>CAMPO</th><th>INCRA</th><th>MEMORIAL</th><th>PROJETO</th><th>OBSERVAÇÃO</th>"
            "\n    </tr>"
            "\n</thead>"
            "\n<tbody>"
            "\n    <tr>"
            "\n        <td><span class='status-alerta'>⚠️</span></td>"
            "\n        <td>[campo]</td>"
            "\n        <td>[valor]</td>"
            "\n        <td>[valor]</td>"
            "\n        <td>[valor]</td>"
            "\n        <td>[explicação]</td>"
            "\n    </tr>"
            "\n</tbody>"
            "\n</table>"
            "\n</div>"
            "\n"
            "\n<!-- SEÇÃO 7: CONSISTÊNCIAS -->"
            "\n<h2>✅ 7. CONSISTÊNCIAS CONFIRMADAS</h2>"
            "\n<div class='secao-ok'>"
            "\n<table>"
            "\n<thead>"
            "\n    <tr style='background:#4caf50;'>"
            "\n        <th>CAMPO</th><th>VALOR CONSISTENTE</th><th>OBSERVAÇÃO</th>"
            "\n    </tr>"
            "\n</thead>"
            "\n<tbody>"
            "\n    <tr>"
            "\n        <td>[campo]</td>"
            "\n        <td>[valor]</td>"
            "\n        <td>Todos os documentos conferem</td>"
            "\n    </tr>"
            "\n</tbody>"
            "\n</table>"
            "\n</div>"
            "\n"
            "\n<!-- SEÇÃO 8: QUALIDADE -->"
            "\n<h2>📝 8. QUALIDADE DOS DOCUMENTOS</h2>"
            "\n<table>"
            "\n<thead>"
            "\n    <tr>"
            "\n        <th>DOCUMENTO</th><th>QUALIDADE</th><th>LEGIBILIDADE</th><th>OBSERVAÇÕES</th>"
            "\n    </tr>"
            "\n</thead>"
            "\n<tbody>"
            "\n    <tr>"
            "\n        <td><strong>INCRA</strong></td>"
            "\n        <td>[Excelente/Boa/Ruim]</td>"
            "\n        <td>[100%/80%/50%]</td>"
            "\n        <td>[comentário]</td>"
            "\n    </tr>"
            "\n    <tr>"
            "\n        <td><strong>MEMORIAL</strong></td>"
            "\n        <td>[Excelente/Boa/Ruim]</td>"
            "\n        <td>[100%/80%/50%]</td>"
            "\n        <td>[comentário]</td>"
            "\n    </tr>"
            "\n    <tr>"
            "\n        <td><strong>PROJETO</strong></td>"
            "\n        <td>[Excelente/Boa/Ruim/N/A]</td>"
            "\n        <td>[100%/80%/50%/N/A]</td>"
            "\n        <td>[comentário]</td>"
            "\n    </tr>"
            "\n</tbody>"
            "\n</table>"
            "\n"
            "\n<!-- SEÇÃO 9: PARECER FINAL -->"
            "\n<h2>⚖️ 9. PARECER FINAL</h2>"
            "\n"
            "\n[Escolha UMA das divs abaixo conforme o resultado:]"
            "\n"
            "\n<div class='parecer parecer-aprovado'>"
            "\n    <p>✅ <strong>APROVADO PARA REGISTRO</strong></p>"
            "\n    <p><strong>Justificativa:</strong> Todos os dados principais estão consistentes entre os documentos.</p>"
            "\n</div>"
            "\n"
            "\n<!-- OU -->"
            "\n"
            "\n<div class='parecer parecer-ressalvas'>"
            "\n    <p>⚠️ <strong>APROVADO COM RESSALVAS</strong></p>"
            "\n    <p><strong>Justificativa:</strong> Há pequenas divergências que não impedem o registro.</p>"
            "\n    <p><strong>Ressalvas:</strong> [listar]</p>"
            "\n</div>"
            "\n"
            "\n<!-- OU -->"
            "\n"
            "\n<div class='parecer parecer-reprovado'>"
            "\n    <p>❌ <strong>REPROVADO - CORREÇÕES OBRIGATÓRIAS</strong></p>"
            "\n    <p><strong>Justificativa:</strong> Discrepâncias críticas impedem o registro.</p>"
            "\n    <p><strong>Correções necessárias:</strong> [listar]</p>"
            "\n</div>"
            "\n"
            "\n<!-- LEGENDA -->"
            "\n<div class='legenda'>"
            "\n    <h3>LEGENDA DE STATUS</h3>"
            "\n    <p><span class='status-ok'>✅</span> = Dados idênticos e corretos</p>"
            "\n    <p><span class='status-alerta'>⚠️</span> = Pequena diferença (revisar, mas não bloqueia)</p>"
            "\n    <p><span class='status-erro'>❌</span> = Erro grave (correção obrigatória)</p>"
            "\n    <p><strong>N/A</strong> = Não encontrado/não aplicável</p>"
            "\n</div>"
            "\n"
            "\n<hr>"
            "\n<p style='text-align:center; color:#888; margin-top:30px;'><em>Relatório gerado por IA - Verificação humana sempre recomendada</em></p>"
            "\n"
            "\n</div>"
            "\n</body>"
            "\n</html>"
            "\n```"
            "\n"
            "\n⚠️ LEMBRE-SE:"
            "\n- Use <span class='status-ok'>✅</span> para dados corretos"
            "\n- Use <span class='status-alerta'>⚠️</span> para pequenas diferenças"
            "\n- Use <span class='status-erro'>❌</span> para erros graves"
            "\n- Escolha APENAS UMA classe de parecer (aprovado/ressalvas/reprovado)"
            "\n- Liste TODOS os vértices encontrados na tabela de coordenadas"
            "\n- Adapte as classes 'resumo' no início conforme o resultado geral"
        )
        
        prompt.append(instrucoes_saida)
        return prompt
        instrucoes_saida = (
            "\n\n"
            "\n════════════════════════════════════════════════════════════════════"
            "\n                    FORMATO DO RELATÓRIO                            "
            "\n════════════════════════════════════════════════════════════════════"
            "\n"
            "\nGere um relatório EXTREMAMENTE ORGANIZADO usando APENAS TABELAS."
            "\nCada tipo de dado deve ter sua própria tabela."
            "\nUse linguagem SIMPLES e DIRETA."
            "\n"
            "\n"
            "\n# 📊 RELATÓRIO DE CONSISTÊNCIA - GEORREFERENCIAMENTO"
            "\n"
            "\n## 🎯 RESUMO EXECUTIVO"
            "\n"
            "\n[Em 2-3 frases diretas: os documentos estão consistentes ou há erros?]"
            "\n"
            "\n---"
            "\n"
            "\n## 📋 1. DADOS CADASTRAIS"
            "\n"
            "\n| DADO | INCRA | MEMORIAL | PROJETO | STATUS |"
            "\n|:-----|:------|:---------|:--------|:------:|"
            "\n| **Proprietário(s)** | [extrair] | [extrair] | [extrair/N/A] | ✅/⚠️/❌ |"
            "\n| **Nome do Imóvel** | [extrair] | [extrair] | [extrair/N/A] | ✅/⚠️/❌ |"
            "\n| **Matrícula(s)** | [extrair] | [extrair] | [extrair/N/A] | ✅/⚠️/❌ |"
            "\n| **Município** | [extrair] | [extrair] | [extrair/N/A] | ✅/⚠️/❌ |"
            "\n| **UF** | [extrair] | [extrair] | [extrair/N/A] | ✅/⚠️/❌ |"
            "\n| **Código INCRA** | [extrair] | [extrair] | [extrair/N/A] | ✅/⚠️/❌ |"
            "\n| **CCIR** | [extrair] | [extrair] | [extrair/N/A] | ✅/⚠️/❌ |"
            "\n"
            "\n**Análise:** [Breve comentário sobre consistência destes dados]"
            "\n"
            "\n---"
            "\n"
            "\n## 📐 2. DADOS TÉCNICOS/MENSURAÇÕES"
            "\n"
            "\n| DADO | INCRA | MEMORIAL | PROJETO | STATUS |"
            "\n|:-----|:------|:---------|:--------|:------:|"
            "\n| **Área Total (ha)** | [X,XXXX] | [X,XXXX] | [X,XXXX/N/A] | ✅/⚠️/❌ |"
            "\n| **Perímetro (m)** | [X.XXX,XX] | [X.XXX,XX] | [X.XXX,XX/N/A] | ✅/⚠️/❌ |"
            "\n| **Sistema Coordenadas** | [UTM/GEO/etc] | [UTM/GEO/etc] | [UTM/GEO/N/A] | ✅/⚠️/❌ |"
            "\n| **Datum** | [SIRGAS/etc] | [SIRGAS/etc] | [SIRGAS/N/A] | ✅/⚠️/❌ |"
            "\n| **Fuso** | [22/23/etc] | [22/23/etc] | [22/23/N/A] | ✅/⚠️/❌ |"
            "\n"
            "\n**Análise:** [Breve comentário sobre consistência destes dados]"
            "\n"
            "\n---"
            "\n"
            "\n## 🗺️ 3. COORDENADAS DOS VÉRTICES"
            "\n"
            "\n**⚠️ CRÍTICO: Liste TODOS os vértices encontrados!**"
            "\n"
            "\n| VÉRTICE | INCRA (Coord) | MEMORIAL (Coord) | PROJETO (Coord) | STATUS |"
            "\n|:--------|:--------------|:-----------------|:----------------|:------:|"
            "\n| **V1** | [E=XXX N=YYY] | [E=XXX N=YYY] | [E=XXX N=YYY/N/A] | ✅/⚠️/❌ |"
            "\n| **V2** | [E=XXX N=YYY] | [E=XXX N=YYY] | [E=XXX N=YYY/N/A] | ✅/⚠️/❌ |"
            "\n| **V3** | [E=XXX N=YYY] | [E=XXX N=YYY] | [E=XXX N=YYY/N/A] | ✅/⚠️/❌ |"
            "\n| **V4** | [E=XXX N=YYY] | [E=XXX N=YYY] | [E=XXX N=YYY/N/A] | ✅/⚠️/❌ |"
            "\n| **V5** | [E=XXX N=YYY] | [E=XXX N=YYY] | [E=XXX N=YYY/N/A] | ✅/⚠️/❌ |"
            "\n| **V6** | [E=XXX N=YYY] | [E=XXX N=YYY] | [E=XXX N=YYY/N/A] | ✅/⚠️/❌ |"
            "\n| **...** | [...] | [...] | [...] | ... |"
            "\n"
            "\n**⚠️ SE HOUVER MAIS VÉRTICES (V7, V8, V9...), ADICIONE MAIS LINHAS!**"
            "\n"
            "\n**Análise:** [Comentário sobre consistência das coordenadas]"
            "\n"
            "\n---"
            "\n"
            "\n## 🧭 4. CONFRONTANTES/LIMITES"
            "\n"
            "\n| DIREÇÃO | INCRA | MEMORIAL | PROJETO | STATUS |"
            "\n|:--------|:------|:---------|:--------|:------:|"
            "\n| **Norte** | [quem/o quê] | [quem/o quê] | [quem/o quê/N/A] | ✅/⚠️/❌ |"
            "\n| **Sul** | [quem/o quê] | [quem/o quê] | [quem/o quê/N/A] | ✅/⚠️/❌ |"
            "\n| **Leste** | [quem/o quê] | [quem/o quê] | [quem/o quê/N/A] | ✅/⚠️/❌ |"
            "\n| **Oeste** | [quem/o quê] | [quem/o quê] | [quem/o quê/N/A] | ✅/⚠️/❌ |"
            "\n"
            "\n**Análise:** [Comentário sobre consistência dos confrontantes]"
            "\n"
            "\n---"
            "\n"
            "\n## 🚨 5. DISCREPÂNCIAS CRÍTICAS"
            "\n"
            "\n[Se NÃO houver erros graves, escreva:]"
            "\n✅ **Nenhuma discrepância crítica identificada.**"
            "\n"
            "\n[Se HOUVER erros graves, use esta tabela:]"
            "\n"
            "\n| TIPO | CAMPO | INCRA | MEMORIAL | PROJETO | AÇÃO NECESSÁRIA |"
            "\n|:-----|:------|:------|:---------|:--------|:----------------|"
            "\n| ❌ | [campo] | [valor] | [valor] | [valor] | [o que corrigir] |"
            "\n| ❌ | [campo] | [valor] | [valor] | [valor] | [o que corrigir] |"
            "\n"
            "\n---"
            "\n"
            "\n## ⚠️ 6. PEQUENAS DIVERGÊNCIAS (Revisar)"
            "\n"
            "\n[Se NÃO houver diferenças pequenas, escreva:]"
            "\n✅ **Nenhuma divergência menor identificada.**"
            "\n"
            "\n[Se HOUVER pequenas diferenças, use esta tabela:]"
            "\n"
            "\n| TIPO | CAMPO | INCRA | MEMORIAL | PROJETO | OBSERVAÇÃO |"
            "\n|:-----|:------|:------|:---------|:--------|:-----------|"
            "\n| ⚠️ | [campo] | [valor] | [valor] | [valor] | [explicação] |"
            "\n| ⚠️ | [campo] | [valor] | [valor] | [valor] | [explicação] |"
            "\n"
            "\n---"
            "\n"
            "\n## ✅ 7. CONSISTÊNCIAS CONFIRMADAS"
            "\n"
            "\n| CAMPO | VALOR CONSISTENTE | OBSERVAÇÃO |"
            "\n|:------|:------------------|:-----------|"
            "\n| [campo] | [valor] | Todos os documentos conferem |"
            "\n| [campo] | [valor] | Todos os documentos conferem |"
            "\n| [campo] | [valor] | Todos os documentos conferem |"
            "\n"
            "\n---"
            "\n"
            "\n## 📝 8. QUALIDADE DOS DOCUMENTOS"
            "\n"
            "\n| DOCUMENTO | QUALIDADE | LEGIBILIDADE | OBSERVAÇÕES |"
            "\n|:----------|:----------|:-------------|:------------|"
            "\n| **INCRA** | [Excelente/Boa/Ruim] | [100%/80%/50%] | [comentário] |"
            "\n| **MEMORIAL** | [Excelente/Boa/Ruim] | [100%/80%/50%] | [comentário] |"
            "\n| **PROJETO** | [Excelente/Boa/Ruim/N/A] | [100%/80%/50%/N/A] | [comentário] |"
            "\n"
            "\n---"
            "\n"
            "\n## ⚖️ 9. PARECER FINAL"
            "\n"
            "\n[Escolha UMA opção e justifique:]"
            "\n"
            "\n### ✅ APROVADO PARA REGISTRO"
            "\n**Justificativa:** Todos os dados principais estão consistentes entre os documentos."
            "\n"
            "\nOU"
            "\n"
            "\n### ⚠️ APROVADO COM RESSALVAS"
            "\n**Justificativa:** Há pequenas divergências que não impedem o registro, mas recomenda-se correção."
            "\n**Ressalvas:** [listar]"
            "\n"
            "\nOU"
            "\n"
            "\n### ❌ REPROVADO - CORREÇÕES OBRIGATÓRIAS"
            "\n**Justificativa:** Discrepâncias críticas impedem o registro."
            "\n**Correções necessárias:** [listar]"
            "\n"
            "\n---"
            "\n"
            "\n**LEGENDA DE STATUS:**"
            "\n- ✅ = Dados idênticos e corretos"
            "\n- ⚠️ = Pequena diferença (revisar, mas não bloqueia)"
            "\n- ❌ = Erro grave (correção obrigatória)"
            "\n- N/A = Não encontrado/não aplicável"
            "\n"
            "\n---"
            "\n*Relatório gerado por IA - Verificação humana sempre recomendada*"
        )
        
        prompt.append(instrucoes_saida)
        return prompt

    def _normalizar_coordenada(self, coord: str) -> str:
        """
        Normaliza coordenadas para comparação, ignorando diferenças de formato.
        Remove "-" do INCRA e "W"/"S" do projeto para comparação equivalente.

        Exemplos:
        - INCRA: "-48°34'14,782"" → "48°34'14,782""
        - PROJETO: "48°34'14,782" W" → "48°34'14,782""
        """
        if not coord:
            return ""

        # Converter para string e remover espaços em branco
        coord = str(coord).strip()

        # Remover "-" do início (INCRA)
        if coord.startswith("-"):
            coord = coord[1:].strip()

        # Remover " W" ou " S" do final (PROJETO)
        coord = coord.replace(" W", "").replace(" S", "").strip()

        # Remover aspas e espaços extras
        coord = coord.strip().strip('"').strip("'").strip()

        return coord

    def _limpar_string(self, valor) -> str:
        """
        Limpa qualquer valor convertendo para string e removendo espaços em branco.
        Remove também caracteres invisíveis que podem causar diferenças falsas.
        Converte pontos decimais em vírgulas para padronização numérica brasileira.
        """
        if valor is None:
            return ""

        # Converter para string e aplicar strip múltiplas vezes
        valor_limpo = str(valor).strip()

        # Remover espaços duplos internos
        while "  " in valor_limpo:
            valor_limpo = valor_limpo.replace("  ", " ")

        # Converter ponto decimal para vírgula (padrão brasileiro)
        valor_limpo = valor_limpo.replace(".", ",")

        return valor_limpo

    def _construir_relatorio_comparacao(self, incluir_projeto: bool, incluir_memorial: bool) -> str:
        """
        Constrói relatório HTML comparando dados estruturados (nova versão V3).
        Compara dados extraídos dos Excel em vez de fazer OCR em tempo real.
        """
        html = []

        # Cabeçalho HTML
        html.append("""<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Relatório de Conferência - Georreferenciamento</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }
        .container {
            max-width: 1400px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        h1 {
            color: #2c3e50;
            text-align: center;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }
        h2 {
            color: #34495e;
            background-color: #ecf0f1;
            padding: 10px;
            border-radius: 5px;
            margin-top: 30px;
        }
        .info-box {
            background-color: #e8f4f8;
            border-left: 4px solid #3498db;
            padding: 15px;
            margin: 20px 0;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }
        th {
            background-color: #3498db;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }
        td {
            padding: 10px;
            border: 1px solid #ddd;
        }
        tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        .identico {
            background-color: #d4edda !important;
        }
        .diferente {
            background-color: #f8d7da !important;
        }
        .status-ok {
            color: #28a745;
            font-weight: bold;
        }
        .status-erro {
            color: #dc3545;
            font-weight: bold;
        }
        .resumo {
            background-color: #fff3cd;
            border: 2px solid #ffc107;
            padding: 20px;
            border-radius: 5px;
            margin: 20px 0;
        }
        .resumo h3 {
            color: #856404;
            margin-top: 0;
        }
        .destaque {
            font-size: 1.1em;
            font-weight: bold;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>📋 RELATÓRIO DE CONFERÊNCIA DE GEORREFERENCIAMENTO</h1>
        <p style="text-align: center; color: #7f8c8d;"><strong>Versão 3.0 - Comparação de Dados Estruturados (Excel)</strong></p>
""")

        # Seção INCRA vs Projeto
        if incluir_projeto and self.projeto_data:
            # Estatísticas
            num_vertices_incra = len(self.incra_data['data'])
            num_vertices_projeto = len(self.projeto_data['data'])

            html.append(f"""
        <div class="info-box">
            <p><strong>📊 Estatísticas:</strong></p>
            <ul>
                <li>Total de vértices INCRA: <strong>{num_vertices_incra}</strong></li>
                <li>Total de vértices PROJETO: <strong>{num_vertices_projeto}</strong></li>
            </ul>
        </div>

        <h2>📐 COMPARAÇÃO: INCRA vs. PROJETO/PLANTA</h2>

        <table>
            <thead>
                <tr>
                    <th style="width: 80px;">Vértice</th>
                    <th style="width: 120px;">Campo</th>
                    <th>INCRA</th>
                    <th>PROJETO</th>
                    <th style="width: 100px;">Status</th>
                </tr>
            </thead>
            <tbody>
""")

            # ===== SEÇÃO 1: COMPARAÇÃO DE VÉRTICE =====
            max_rows = max(num_vertices_incra, num_vertices_projeto)
            diferencas_vertice = 0
            identicos_vertice = 0
            diferencas_segmento = 0
            identicos_segmento = 0

            for i in range(max_rows):
                incra_row = self.incra_data['data'][i] if i < num_vertices_incra else None
                projeto_row = self.projeto_data['data'][i] if i < num_vertices_projeto else None

                if incra_row and projeto_row:
                    # Extrair e limpar dados VÉRTICE (colunas 0-3)
                    codigo_incra = self._limpar_string(incra_row[0] if len(incra_row) > 0 else "")
                    codigo_projeto = self._limpar_string(projeto_row[0] if len(projeto_row) > 0 else "")

                    long_incra = self._limpar_string(incra_row[1] if len(incra_row) > 1 else "")
                    long_projeto = self._limpar_string(projeto_row[1] if len(projeto_row) > 1 else "")

                    lat_incra = self._limpar_string(incra_row[2] if len(incra_row) > 2 else "")
                    lat_projeto = self._limpar_string(projeto_row[2] if len(projeto_row) > 2 else "")

                    alt_incra = self._limpar_string(incra_row[3] if len(incra_row) > 3 else "")
                    alt_projeto = self._limpar_string(projeto_row[3] if len(projeto_row) > 3 else "")

                    # Normalizar coordenadas para comparação
                    long_incra_norm = self._normalizar_coordenada(long_incra)
                    long_projeto_norm = self._normalizar_coordenada(long_projeto)

                    lat_incra_norm = self._normalizar_coordenada(lat_incra)
                    lat_projeto_norm = self._normalizar_coordenada(lat_projeto)

                    # Verificar se VÉRTICE é idêntico (comparando strings limpas)
                    vertice_identico = (codigo_incra == codigo_projeto and
                                       long_incra_norm == long_projeto_norm and
                                       lat_incra_norm == lat_projeto_norm and
                                       alt_incra == alt_projeto)

                    if vertice_identico:
                        status_class_vertice = "identico"
                        status_texto_vertice = '<span class="status-ok">✅ IDÊNTICO</span>'
                        identicos_vertice += 1
                    else:
                        status_class_vertice = "diferente"
                        status_texto_vertice = '<span class="status-erro">❌ DIFERENTE</span>'
                        diferencas_vertice += 1

                    # Adicionar linhas VÉRTICE na tabela
                    html.append(f"""
                <tr class="{status_class_vertice}">
                    <td rowspan="4" style="text-align: center; vertical-align: middle; font-weight: bold;">#{i+1}</td>
                    <td><strong>Código</strong></td>
                    <td>{codigo_incra}</td>
                    <td>{codigo_projeto}</td>
                    <td rowspan="4" style="text-align: center; vertical-align: middle;">{status_texto_vertice}</td>
                </tr>
                <tr class="{status_class_vertice}">
                    <td><strong>Longitude</strong></td>
                    <td>{long_incra}</td>
                    <td>{long_projeto}</td>
                </tr>
                <tr class="{status_class_vertice}">
                    <td><strong>Latitude</strong></td>
                    <td>{lat_incra}</td>
                    <td>{lat_projeto}</td>
                </tr>
                <tr class="{status_class_vertice}">
                    <td><strong>Altitude</strong></td>
                    <td>{alt_incra}</td>
                    <td>{alt_projeto}</td>
                </tr>
""")

                elif incra_row and not projeto_row:
                    diferencas_vertice += 1
                    html.append(f"""
                <tr class="diferente">
                    <td style="text-align: center; font-weight: bold;">#{i+1}</td>
                    <td colspan="3"><strong>❌ AUSENTE NO PROJETO</strong> - Código INCRA: {incra_row[0]}</td>
                    <td style="text-align: center;"><span class="status-erro">❌ ERRO</span></td>
                </tr>
""")

                elif not incra_row and projeto_row:
                    diferencas_vertice += 1
                    html.append(f"""
                <tr class="diferente">
                    <td style="text-align: center; font-weight: bold;">#{i+1}</td>
                    <td colspan="3"><strong>❌ EXTRA NO PROJETO</strong> (não existe no INCRA) - Código: {projeto_row[0]}</td>
                    <td style="text-align: center;"><span class="status-erro">❌ ERRO</span></td>
                </tr>
""")

            html.append("""
            </tbody>
        </table>
""")

            # ===== SEÇÃO 2: COMPARAÇÃO DE SEGMENTO VANTE =====
            html.append("""
        <h2>🔄 COMPARAÇÃO: SEGMENTO VANTE</h2>

        <table>
            <thead>
                <tr>
                    <th style="width: 80px;">Vértice</th>
                    <th style="width: 120px;">Campo</th>
                    <th>INCRA</th>
                    <th>PROJETO</th>
                    <th style="width: 100px;">Status</th>
                </tr>
            </thead>
            <tbody>
""")

            for i in range(max_rows):
                incra_row = self.incra_data['data'][i] if i < num_vertices_incra else None
                projeto_row = self.projeto_data['data'][i] if i < num_vertices_projeto else None

                if incra_row and projeto_row:
                    # Extrair e limpar dados SEGMENTO VANTE (colunas 4-6)
                    cod_seg_incra = self._limpar_string(incra_row[4] if len(incra_row) > 4 else "")
                    cod_seg_projeto = self._limpar_string(projeto_row[4] if len(projeto_row) > 4 else "")

                    azim_incra = self._limpar_string(incra_row[5] if len(incra_row) > 5 else "")
                    azim_projeto = self._limpar_string(projeto_row[5] if len(projeto_row) > 5 else "")

                    dist_incra = self._limpar_string(incra_row[6] if len(incra_row) > 6 else "")
                    dist_projeto = self._limpar_string(projeto_row[6] if len(projeto_row) > 6 else "")

                    # Verificar se SEGMENTO VANTE é idêntico (comparando strings limpas)
                    segmento_identico = (cod_seg_incra == cod_seg_projeto and
                                        azim_incra == azim_projeto and
                                        dist_incra == dist_projeto)

                    if segmento_identico:
                        status_class_seg = "identico"
                        status_texto_seg = '<span class="status-ok">✅ IDÊNTICO</span>'
                        identicos_segmento += 1
                    else:
                        status_class_seg = "diferente"
                        status_texto_seg = '<span class="status-erro">❌ DIFERENTE</span>'
                        diferencas_segmento += 1

                    # Adicionar linhas SEGMENTO VANTE na tabela
                    html.append(f"""
                <tr class="{status_class_seg}">
                    <td rowspan="3" style="text-align: center; vertical-align: middle; font-weight: bold;">#{i+1}</td>
                    <td><strong>Código</strong></td>
                    <td>{cod_seg_incra}</td>
                    <td>{cod_seg_projeto}</td>
                    <td rowspan="3" style="text-align: center; vertical-align: middle;">{status_texto_seg}</td>
                </tr>
                <tr class="{status_class_seg}">
                    <td><strong>Azimute</strong></td>
                    <td>{azim_incra}</td>
                    <td>{azim_projeto}</td>
                </tr>
                <tr class="{status_class_seg}">
                    <td><strong>Dist. (m)</strong></td>
                    <td>{dist_incra}</td>
                    <td>{dist_projeto}</td>
                </tr>
""")

                elif incra_row and not projeto_row:
                    diferencas_segmento += 1
                    html.append(f"""
                <tr class="diferente">
                    <td style="text-align: center; font-weight: bold;">#{i+1}</td>
                    <td colspan="3"><strong>❌ AUSENTE NO PROJETO</strong></td>
                    <td style="text-align: center;"><span class="status-erro">❌ ERRO</span></td>
                </tr>
""")

                elif not incra_row and projeto_row:
                    diferencas_segmento += 1
                    html.append(f"""
                <tr class="diferente">
                    <td style="text-align: center; font-weight: bold;">#{i+1}</td>
                    <td colspan="3"><strong>❌ EXTRA NO PROJETO</strong></td>
                    <td style="text-align: center;"><span class="status-erro">❌ ERRO</span></td>
                </tr>
""")

            html.append("""
            </tbody>
        </table>
""")

            # Resumo geral
            diferencas_total = diferencas_vertice + diferencas_segmento
            identicos_total = identicos_vertice + identicos_segmento
            resultado_final = "🎉 TODOS OS DADOS ESTÃO IDÊNTICOS!" if diferencas_total == 0 else "⚠️ EXISTEM DIFERENÇAS ENTRE OS DOCUMENTOS"
            resultado_cor = "#28a745" if diferencas_total == 0 else "#dc3545"

            html.append(f"""
        <div class="resumo">
            <h3>📊 RESUMO DA COMPARAÇÃO</h3>
            <p class="destaque">Total de vértices analisados: {max_rows}</p>

            <h4 style="margin-top: 20px; color: #2c3e50;">📍 VÉRTICE (Código, Longitude, Latitude, Altitude):</h4>
            <p>✅ Idênticos: <strong style="color: #28a745;">{identicos_vertice}</strong></p>
            <p>❌ Diferentes: <strong style="color: #dc3545;">{diferencas_vertice}</strong></p>

            <h4 style="margin-top: 20px; color: #2c3e50;">🔄 SEGMENTO VANTE (Código, Azimute, Distância):</h4>
            <p>✅ Idênticos: <strong style="color: #28a745;">{identicos_segmento}</strong></p>
            <p>❌ Diferentes: <strong style="color: #dc3545;">{diferencas_segmento}</strong></p>

            <hr style="margin: 20px 0;">

            <h4 style="color: #2c3e50;">🎯 TOTAL GERAL:</h4>
            <p>✅ Total idênticos: <strong style="color: #28a745;">{identicos_total}</strong></p>
            <p>❌ Total diferentes: <strong style="color: #dc3545;">{diferencas_total}</strong></p>

            <hr style="margin: 20px 0;">
            <p class="destaque" style="color: {resultado_cor}; font-size: 1.2em;">{resultado_final}</p>
            {f'<p style="color: #856404;">Por favor, revise os itens marcados como DIFERENTE nas tabelas acima.</p>' if diferencas_total > 0 else ''}
        </div>
""")

        # Informações do processo
        html.append(f"""
        <div class="info-box">
            <h3>📁 INFORMAÇÕES DO PROCESSO</h3>
            <p><strong>Arquivos Excel gerados para auditoria:</strong></p>
            <ul>
                <li>INCRA: <code>{self.incra_excel_path}</code></li>
                <li>PROJETO: <code>{self.projeto_excel_path}</code></li>
            </ul>
        </div>

        <p style="text-align: center; color: #7f8c8d; margin-top: 40px;">
            <em>Relatório gerado automaticamente - Versão 3.0</em>
        </p>
    </div>
</body>
</html>
""")

        return "".join(html)

    def _executar_analise_gemini(self, incluir_projeto: bool = False, incluir_memorial: bool = False):
        """
        Executa a análise completa usando extração para Excel + comparação.
        Nova versão V3: Extrai PDFs para Excel primeiro, depois compara dados estruturados.
        Deve ser executado em thread separada para não travar a GUI.
        """
        try:
            # Limpar área de resultados
            self.resultado_text.delete(1.0, tk.END)
            self.resultado_text.insert(tk.END, "🔄 Processando documentos com NOVA ABORDAGEM V3...\n\n")
            self.resultado_text.insert(tk.END, "📊 Fluxo: PDF → Extração para Excel → Comparação de dados estruturados\n\n")
            self.resultado_text.insert(tk.END, "="*80 + "\n\n")

            # ===== ETAPA 1: EXTRAIR INCRA PARA EXCEL =====
            self._atualizar_status("Extraindo tabela do INCRA para Excel...")
            self.resultado_text.insert(tk.END, "🔄 [1/2] Extraindo INCRA para Excel...\n")
            self.resultado_text.insert(tk.END, f"    PDF: {self.incra_path.get()}\n")
            self.root.update_idletasks()

            try:
                self.incra_excel_path, self.incra_data = self._extrair_pdf_para_excel(
                    self.incra_path.get(),
                    tipo="incra"
                )
                self.resultado_text.insert(
                    tk.END,
                    f"✅ INCRA extraído com sucesso!\n"
                    f"    Vértices: {len(self.incra_data['data'])}\n"
                    f"    Excel: {self.incra_excel_path}\n\n"
                )
                self.root.update_idletasks()
            except Exception as e:
                raise RuntimeError(f"Erro ao extrair INCRA: {str(e)}") from e

            # ===== ETAPA 2: EXTRAIR PROJETO PARA EXCEL =====
            self._atualizar_status("Extraindo tabela do Projeto para Excel...")
            self.resultado_text.insert(tk.END, "🔄 [2/2] Extraindo Projeto para Excel...\n")
            self.resultado_text.insert(tk.END, f"    PDF: {self.projeto_path.get()}\n")
            self.root.update_idletasks()

            try:
                self.projeto_excel_path, self.projeto_data = self._extrair_pdf_para_excel(
                    self.projeto_path.get(),
                    tipo="normal"
                )
                self.resultado_text.insert(
                    tk.END,
                    f"✅ Projeto extraído com sucesso!\n"
                    f"    Vértices: {len(self.projeto_data['data'])}\n"
                    f"    Excel: {self.projeto_excel_path}\n\n"
                )
                self.root.update_idletasks()
            except Exception as e:
                raise RuntimeError(f"Erro ao extrair PROJETO: {str(e)}") from e

            self.resultado_text.insert(tk.END, "="*80 + "\n\n")

            # ===== ETAPA 3: COMPARAR DADOS ESTRUTURADOS =====
            self._atualizar_status("Comparando dados estruturados...")
            self.resultado_text.insert(tk.END, "🔄 Comparando dados estruturados...\n\n")
            self.root.update_idletasks()

            # Construir relatório de comparação HTML
            relatorio_html = self._construir_relatorio_comparacao(True, False)

            # Salvar HTML automaticamente
            output_dir = Path(tempfile.gettempdir()) / "conferencia_geo"
            output_dir.mkdir(exist_ok=True)
            html_path = output_dir / "relatorio_comparacao.html"

            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(relatorio_html)

            # Salvar HTML para exportação futura
            self.ultimo_relatorio_html = relatorio_html

            # Exibir resumo no ScrolledText
            self.resultado_text.insert(tk.END, "="*80 + "\n")
            self.resultado_text.insert(tk.END, "✅ ANÁLISE CONCLUÍDA COM SUCESSO!\n")
            self.resultado_text.insert(tk.END, "="*80 + "\n\n")

            # Contar diferenças para o resumo
            num_vertices = len(self.incra_data['data'])
            self.resultado_text.insert(tk.END, f"📊 Total de vértices analisados: {num_vertices}\n\n")

            self.resultado_text.insert(tk.END, "📁 ARQUIVOS GERADOS:\n")
            self.resultado_text.insert(tk.END, f"   • INCRA (Excel): {self.incra_excel_path}\n")
            self.resultado_text.insert(tk.END, f"   • PROJETO (Excel): {self.projeto_excel_path}\n")
            self.resultado_text.insert(tk.END, f"   • RELATÓRIO (HTML): {html_path}\n\n")

            self.resultado_text.insert(tk.END, "="*80 + "\n")
            self.resultado_text.insert(tk.END, "🌐 O relatório HTML foi aberto automaticamente no navegador!\n")
            self.resultado_text.insert(tk.END, "="*80 + "\n")

            # Habilitar botão de salvar
            self.btn_salvar_html.config(state='normal')

            self._atualizar_status("✅ Análise concluída!")

            # Abrir HTML no navegador automaticamente
            import webbrowser
            webbrowser.open(f'file://{html_path}')

            messagebox.showinfo("Sucesso",
                              f"Análise concluída com sucesso!\n\n"
                              f"✅ Dados extraídos para Excel\n"
                              f"✅ Comparação estruturada realizada\n"
                              f"✅ Relatório HTML aberto no navegador\n\n"
                              f"Arquivo: {html_path}")

        except Exception as e:
            import traceback
            import sys

            # Capturar traceback completo
            tb_str = traceback.format_exc()

            # Mostrar erro detalhado na GUI
            erro_msg = f"\n\n{'='*80}\n❌ ERRO DURANTE A ANÁLISE\n{'='*80}\n\n"
            erro_msg += f"Tipo: {type(e).__name__}\n"
            erro_msg += f"Mensagem: {str(e)}\n\n"
            erro_msg += "Detalhes técnicos:\n"
            erro_msg += "-" * 80 + "\n"
            erro_msg += tb_str
            erro_msg += "-" * 80 + "\n\n"
            erro_msg += "💡 Dicas para resolver:\n"
            erro_msg += "- Verifique se os arquivos PDF estão acessíveis\n"
            erro_msg += "- Verifique se você tem permissão para criar arquivos em:\n"
            erro_msg += f"  {Path(tempfile.gettempdir()) / 'conferencia_geo'}\n"
            erro_msg += "- Verifique sua conexão com a API do Gemini\n"
            erro_msg += "- Tente fechar outros programas que possam estar usando os arquivos\n"

            self.resultado_text.insert(tk.END, erro_msg)
            self._atualizar_status("❌ Erro na análise")

            # Mostrar erro em popup simplificado
            messagebox.showerror("Erro na Análise",
                               f"Ocorreu um erro durante a análise:\n\n"
                               f"{type(e).__name__}: {str(e)}\n\n"
                               f"Veja detalhes completos na área de resultados.")

            # Também imprimir no console para debug
            print(erro_msg, file=sys.stderr)

        finally:
            self._habilitar_botoes()

    def _comparar_projeto(self):
        """Compara INCRA vs. Projeto."""
        if not self._validar_entrada():
            return

        self._desabilitar_botoes()

        # Executar em thread separada para não travar a GUI
        thread = threading.Thread(target=self._executar_analise_gemini, args=(True, False))
        thread.daemon = True
        thread.start()


class JanelaComparacaoManual:
    """Janela para comparação visual manual dos documentos PDF."""
    
    def __init__(self, parent, incra_path, memorial_path, projeto_path=None):
        self.janela = tk.Toplevel(parent)
        self.janela.title("Comparação Visual Manual - Georreferenciamento")
        self.janela.geometry("1600x900")
        self.janela.configure(bg='#2c3e50')
        
        # Caminhos dos arquivos
        self.incra_path = incra_path
        self.memorial_path = memorial_path
        self.projeto_path = projeto_path
        
        # Listas de imagens carregadas
        self.incra_images = []
        self.memorial_images = []
        self.projeto_images = []
        
        # Índices de página atual
        self.incra_pagina = 0
        self.memorial_pagina = 0
        self.projeto_pagina = 0
        
        # Níveis de zoom (100% = 1.0)
        self.incra_zoom = 1.0
        self.memorial_zoom = 1.0
        self.projeto_zoom = 1.0
        
        # Ângulo de rotação (0, 90, 180, 270)
        self.incra_rotacao = 0
        self.memorial_rotacao = 0
        self.projeto_rotacao = 0
        
        # Posição do canvas (para arrastar)
        self.incra_pos_x = 0
        self.incra_pos_y = 0
        self.memorial_pos_x = 0
        self.memorial_pos_y = 0
        self.projeto_pos_x = 0
        self.projeto_pos_y = 0
        
        # Controle de arrastar
        self.incra_drag_start = None
        self.memorial_drag_start = None
        self.projeto_drag_start = None
        
        # Imagens PhotoImage (para exibição no Tkinter)
        self.incra_photo = None
        self.memorial_photo = None
        self.projeto_photo = None
        
        self._criar_interface()
        self._carregar_documentos()
        
    def _criar_interface(self):
        """Cria a interface da janela de comparação."""
        
        # Frame superior com título
        header_frame = tk.Frame(self.janela, bg='#34495e', height=60)
        header_frame.pack(fill=tk.X, side=tk.TOP)
        header_frame.pack_propagate(False)
        
        titulo = tk.Label(
            header_frame,
            text="👁️  COMPARAÇÃO VISUAL MANUAL",
            font=('Arial', 18, 'bold'),
            bg='#34495e',
            fg='white'
        )
        titulo.pack(pady=15)
        
        # Frame principal com painéis
        main_frame = tk.Frame(self.janela, bg='#2c3e50')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Determinar quantos painéis criar
        num_paineis = 3 if self.projeto_path else 2
        
        # Criar painéis lado a lado
        if num_paineis == 2:
            # INCRA e Memorial
            self._criar_painel(main_frame, "INCRA", 0, 'incra')
            self._criar_painel(main_frame, "MEMORIAL", 1, 'memorial')
        else:
            # INCRA, Memorial e Projeto
            self._criar_painel(main_frame, "INCRA", 0, 'incra', largura_col=3)
            self._criar_painel(main_frame, "MEMORIAL", 1, 'memorial', largura_col=3)
            self._criar_painel(main_frame, "PROJETO", 2, 'projeto', largura_col=3)
        
        # Frame inferior com instruções
        footer_frame = tk.Frame(self.janela, bg='#34495e', height=50)
        footer_frame.pack(fill=tk.X, side=tk.BOTTOM)
        footer_frame.pack_propagate(False)
        
        instrucoes = tk.Label(
            footer_frame,
            text="💡 Zoom: +/- ou Scroll Mouse | Páginas: ◀️ ▶️ | Girar: 🔄 90° | Arrastar: Segurar botão esquerdo",
            font=('Arial', 10),
            bg='#34495e',
            fg='#ecf0f1'
        )
        instrucoes.pack(pady=12)
        
    def _criar_painel(self, parent, titulo, coluna, tipo, largura_col=2):
        """Cria um painel de visualização para um documento."""
        
        # Frame do painel
        painel = tk.Frame(parent, bg='#ecf0f1', relief=tk.RAISED, borderwidth=2)
        painel.grid(row=0, column=coluna, sticky=(tk.N, tk.S, tk.E, tk.W), padx=5, pady=5)
        
        parent.columnconfigure(coluna, weight=1)
        parent.rowconfigure(0, weight=1)
        
        # Cabeçalho do painel
        header = tk.Frame(painel, bg='#3498db', height=50)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        tk.Label(
            header,
            text=f"📄 {titulo}",
            font=('Arial', 14, 'bold'),
            bg='#3498db',
            fg='white'
        ).pack(pady=10)
        
        # Frame para canvas com scrollbar
        canvas_frame = tk.Frame(painel, bg='white')
        canvas_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Canvas para imagem
        canvas = tk.Canvas(canvas_frame, bg='white', highlightthickness=0)
        canvas.pack(fill=tk.BOTH, expand=True)
        
        # Salvar referência ao canvas
        setattr(self, f'{tipo}_canvas', canvas)
        
        # Configurar eventos do mouse para arrastar e zoom
        canvas.bind('<ButtonPress-1>', lambda e: self._iniciar_arrasto(tipo, e))
        canvas.bind('<B1-Motion>', lambda e: self._arrastar(tipo, e))
        canvas.bind('<ButtonRelease-1>', lambda e: self._finalizar_arrasto(tipo))
        canvas.bind('<MouseWheel>', lambda e: self._zoom_scroll(tipo, e))
        # Para Linux
        canvas.bind('<Button-4>', lambda e: self._zoom_scroll(tipo, e))
        canvas.bind('<Button-5>', lambda e: self._zoom_scroll(tipo, e))
        
        # Frame de controles
        controles = tk.Frame(painel, bg='#ecf0f1', height=120)
        controles.pack(fill=tk.X)
        controles.pack_propagate(False)
        
        # Linha 1: Navegação de páginas
        nav_frame = tk.Frame(controles, bg='#ecf0f1')
        nav_frame.pack(pady=5)
        
        btn_anterior = tk.Button(
            nav_frame,
            text="◀️ Anterior",
            command=lambda: self._mudar_pagina(tipo, -1),
            font=('Arial', 10),
            bg='#95a5a6',
            fg='white',
            relief=tk.RAISED,
            padx=10,
            pady=5
        )
        btn_anterior.pack(side=tk.LEFT, padx=5)
        
        # Label da página atual
        label_pagina = tk.Label(
            nav_frame,
            text="Página 1/1",
            font=('Arial', 10, 'bold'),
            bg='#ecf0f1'
        )
        label_pagina.pack(side=tk.LEFT, padx=10)
        setattr(self, f'{tipo}_label_pagina', label_pagina)
        
        btn_proximo = tk.Button(
            nav_frame,
            text="Próxima ▶️",
            command=lambda: self._mudar_pagina(tipo, 1),
            font=('Arial', 10),
            bg='#95a5a6',
            fg='white',
            relief=tk.RAISED,
            padx=10,
            pady=5
        )
        btn_proximo.pack(side=tk.LEFT, padx=5)
        
        # Linha 2: Controles de zoom
        zoom_frame = tk.Frame(controles, bg='#ecf0f1')
        zoom_frame.pack(pady=5)
        
        btn_zoom_out = tk.Button(
            zoom_frame,
            text="➖ Zoom -",
            command=lambda: self._ajustar_zoom(tipo, -0.2),
            font=('Arial', 10),
            bg='#e74c3c',
            fg='white',
            relief=tk.RAISED,
            padx=10,
            pady=5
        )
        btn_zoom_out.pack(side=tk.LEFT, padx=5)
        
        btn_zoom_reset = tk.Button(
            zoom_frame,
            text="🔄 Reset",
            command=lambda: self._resetar_zoom(tipo),
            font=('Arial', 10),
            bg='#95a5a6',
            fg='white',
            relief=tk.RAISED,
            padx=10,
            pady=5
        )
        btn_zoom_reset.pack(side=tk.LEFT, padx=5)
        
        btn_zoom_in = tk.Button(
            zoom_frame,
            text="➕ Zoom +",
            command=lambda: self._ajustar_zoom(tipo, 0.2),
            font=('Arial', 10),
            bg='#27ae60',
            fg='white',
            relief=tk.RAISED,
            padx=10,
            pady=5
        )
        btn_zoom_in.pack(side=tk.LEFT, padx=5)
        
        # Label do zoom atual
        label_zoom = tk.Label(
            zoom_frame,
            text="100%",
            font=('Arial', 10),
            bg='#ecf0f1'
        )
        label_zoom.pack(side=tk.LEFT, padx=10)
        setattr(self, f'{tipo}_label_zoom', label_zoom)
        
        # Linha 3: Controles de rotação
        rotacao_frame = tk.Frame(controles, bg='#ecf0f1')
        rotacao_frame.pack(pady=5)
        
        btn_girar = tk.Button(
            rotacao_frame,
            text="🔄 Girar 90°",
            command=lambda: self._girar_imagem(tipo),
            font=('Arial', 10),
            bg='#3498db',
            fg='white',
            relief=tk.RAISED,
            padx=10,
            pady=5
        )
        btn_girar.pack(side=tk.LEFT, padx=5)
        
        btn_resetar_rotacao = tk.Button(
            rotacao_frame,
            text="↻ Reset Rotação",
            command=lambda: self._resetar_rotacao(tipo),
            font=('Arial', 10),
            bg='#9b59b6',
            fg='white',
            relief=tk.RAISED,
            padx=10,
            pady=5
        )
        btn_resetar_rotacao.pack(side=tk.LEFT, padx=5)
        
        # Label da rotação atual
        label_rotacao = tk.Label(
            rotacao_frame,
            text="0°",
            font=('Arial', 10),
            bg='#ecf0f1'
        )
        label_rotacao.pack(side=tk.LEFT, padx=10)
        setattr(self, f'{tipo}_label_rotacao', label_rotacao)
        
    def _carregar_documentos(self):
        """Carrega os documentos PDF como imagens."""
        try:
            # Criar diálogo de progresso
            progress = tk.Toplevel(self.janela)
            progress.title("Carregando...")
            progress.geometry("400x150")
            progress.transient(self.janela)
            progress.grab_set()
            
            tk.Label(
                progress,
                text="⏳ Carregando documentos...",
                font=('Arial', 12, 'bold')
            ).pack(pady=20)
            
            status_label = tk.Label(progress, text="", font=('Arial', 10))
            status_label.pack(pady=10)
            
            progress.update()
            
            # Carregar INCRA (com rotação)
            status_label.config(text="Carregando INCRA...")
            progress.update()
            self.incra_images = convert_from_path(self.incra_path, dpi=150)
            # Rotacionar INCRA
            self.incra_images = [img.rotate(-90, expand=True) for img in self.incra_images]
            
            # Carregar Memorial
            status_label.config(text="Carregando Memorial...")
            progress.update()
            self.memorial_images = convert_from_path(self.memorial_path, dpi=150)
            
            # Carregar Projeto se houver
            if self.projeto_path:
                status_label.config(text="Carregando Projeto...")
                progress.update()
                self.projeto_images = convert_from_path(self.projeto_path, dpi=150)
            
            progress.destroy()
            
            # Exibir primeira página de cada documento
            self._exibir_pagina('incra')
            self._exibir_pagina('memorial')
            if self.projeto_path:
                self._exibir_pagina('projeto')
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar documentos:\n{str(e)}")
            self.janela.destroy()
            
    def _exibir_pagina(self, tipo):
        """Exibe a página atual de um documento."""
        # Obter lista de imagens e índice atual
        images = getattr(self, f'{tipo}_images')
        pagina = getattr(self, f'{tipo}_pagina')
        zoom = getattr(self, f'{tipo}_zoom')
        rotacao = getattr(self, f'{tipo}_rotacao')
        pos_x = getattr(self, f'{tipo}_pos_x')
        pos_y = getattr(self, f'{tipo}_pos_y')
        canvas = getattr(self, f'{tipo}_canvas')
        
        if not images or pagina >= len(images):
            return
            
        # Obter imagem original
        img_original = images[pagina].copy()
        
        # Aplicar rotação (se houver)
        if rotacao != 0:
            img_original = img_original.rotate(-rotacao, expand=True)
        
        # Aplicar zoom
        largura = int(img_original.width * zoom)
        altura = int(img_original.height * zoom)
        img_zoom = img_original.resize((largura, altura), Image.Resampling.LANCZOS)
        
        # Converter para PhotoImage
        photo = ImageTk.PhotoImage(img_zoom)
        setattr(self, f'{tipo}_photo', photo)  # Manter referência
        
        # Limpar canvas e exibir imagem
        canvas.delete("all")
        canvas.create_image(pos_x, pos_y, anchor=tk.NW, image=photo, tags='imagem')
        canvas.config(scrollregion=canvas.bbox("all"))
        
        # Atualizar label de página
        label_pagina = getattr(self, f'{tipo}_label_pagina')
        label_pagina.config(text=f"Página {pagina + 1}/{len(images)}")
        
        # Atualizar label de zoom
        label_zoom = getattr(self, f'{tipo}_label_zoom')
        label_zoom.config(text=f"{int(zoom * 100)}%")
        
        # Atualizar label de rotação
        label_rotacao = getattr(self, f'{tipo}_label_rotacao')
        label_rotacao.config(text=f"{rotacao}°")
        
    def _mudar_pagina(self, tipo, direcao):
        """Muda para página anterior ou próxima."""
        images = getattr(self, f'{tipo}_images')
        pagina_atual = getattr(self, f'{tipo}_pagina')
        
        nova_pagina = pagina_atual + direcao
        
        # Verificar limites
        if 0 <= nova_pagina < len(images):
            setattr(self, f'{tipo}_pagina', nova_pagina)
            self._exibir_pagina(tipo)
            
    def _ajustar_zoom(self, tipo, delta):
        """Ajusta o nível de zoom."""
        zoom_atual = getattr(self, f'{tipo}_zoom')
        novo_zoom = max(0.2, min(3.0, zoom_atual + delta))  # Limitar entre 20% e 300%
        
        setattr(self, f'{tipo}_zoom', novo_zoom)
        self._exibir_pagina(tipo)
        
    def _resetar_zoom(self, tipo):
        """Reseta o zoom para 100%."""
        setattr(self, f'{tipo}_zoom', 1.0)
        self._exibir_pagina(tipo)
    
    def _girar_imagem(self, tipo):
        """Gira a imagem em 90 graus no sentido horário."""
        rotacao_atual = getattr(self, f'{tipo}_rotacao')
        nova_rotacao = (rotacao_atual + 90) % 360
        setattr(self, f'{tipo}_rotacao', nova_rotacao)
        
        # Resetar posição ao girar
        setattr(self, f'{tipo}_pos_x', 0)
        setattr(self, f'{tipo}_pos_y', 0)
        
        self._exibir_pagina(tipo)
    
    def _resetar_rotacao(self, tipo):
        """Reseta a rotação para 0 graus."""
        setattr(self, f'{tipo}_rotacao', 0)
        setattr(self, f'{tipo}_pos_x', 0)
        setattr(self, f'{tipo}_pos_y', 0)
        self._exibir_pagina(tipo)
    
    def _iniciar_arrasto(self, tipo, event):
        """Inicia o arrasto da imagem."""
        canvas = getattr(self, f'{tipo}_canvas')
        canvas.config(cursor="fleur")  # Cursor de mover
        setattr(self, f'{tipo}_drag_start', (event.x, event.y))
    
    def _arrastar(self, tipo, event):
        """Arrasta a imagem."""
        drag_start = getattr(self, f'{tipo}_drag_start')
        if drag_start is None:
            return
        
        # Calcular deslocamento
        dx = event.x - drag_start[0]
        dy = event.y - drag_start[1]
        
        # Atualizar posição
        pos_x = getattr(self, f'{tipo}_pos_x')
        pos_y = getattr(self, f'{tipo}_pos_y')
        
        setattr(self, f'{tipo}_pos_x', pos_x + dx)
        setattr(self, f'{tipo}_pos_y', pos_y + dy)
        
        # Atualizar ponto de início
        setattr(self, f'{tipo}_drag_start', (event.x, event.y))
        
        # Redesenhar
        self._exibir_pagina(tipo)
    
    def _finalizar_arrasto(self, tipo):
        """Finaliza o arrasto da imagem."""
        canvas = getattr(self, f'{tipo}_canvas')
        canvas.config(cursor="")  # Cursor normal
        setattr(self, f'{tipo}_drag_start', None)
    
    def _zoom_scroll(self, tipo, event):
        """Ajusta o zoom com o scroll do mouse."""
        # Determinar direção do scroll
        if event.num == 4 or event.delta > 0:
            # Scroll para cima = zoom in
            delta = 0.1
        elif event.num == 5 or event.delta < 0:
            # Scroll para baixo = zoom out
            delta = -0.1
        else:
            return
        
        # Ajustar zoom
        zoom_atual = getattr(self, f'{tipo}_zoom')
        novo_zoom = max(0.2, min(5.0, zoom_atual + delta))  # Limitar entre 20% e 500%
        
        setattr(self, f'{tipo}_zoom', novo_zoom)
        self._exibir_pagina(tipo)


def main():
    """Função principal para iniciar a aplicação."""
    root = tk.Tk()
    app = VerificadorGeorreferenciamento(root)
    root.mainloop()


if __name__ == "__main__":
    main()