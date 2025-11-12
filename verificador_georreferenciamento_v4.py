#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verificador de Consistência de Documentos de Georreferenciamento
Aplicação GUI para cartórios - Análise multimodal com Gemini AI
Autor: Sistema Automatizado
Versão: 4.0 - Interface moderna com Modo Automático
"""

import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from tkinter import ttk
from pathlib import Path
import threading
from typing import List, Optional, Dict, Tuple
import json
import tempfile
import shutil
import webbrowser
import math
from datetime import datetime
import configparser

try:
    from pdf2image import convert_from_path
    from PIL import Image, ImageTk
    import google.generativeai as genai
    from openpyxl import load_workbook
    import PyPDF2
    # Importar funções de extração do script existente
    from process_memorial_descritivo_v2 import (
        extract_table_from_pdf,
        extrair_memorial_incra,
        create_excel_file
    )
except ImportError as e:
    print(f"❌ Erro: Biblioteca necessária não encontrada: {e}")
    print("\nInstale as dependências com:")
    print("pip install pdf2image Pillow google-generativeai openpyxl PyPDF2")
    print("\nNota: Também é necessário ter o 'poppler-utils' instalado no sistema.")
    sys.exit(1)


class ConfigManager:
    """Gerencia configurações persistentes da aplicação."""

    def __init__(self):
        self.config_dir = Path.home() / ".conferencia_geo"
        self.config_file = self.config_dir / "config.ini"
        self.config = configparser.ConfigParser()
        self._ensure_config_exists()

    def _ensure_config_exists(self):
        """Cria diretório e arquivo de configuração se não existir."""
        self.config_dir.mkdir(parents=True, exist_ok=True)
        if not self.config_file.exists():
            self.config['API'] = {'gemini_key': ''}
            self.save()
        else:
            self.config.read(self.config_file)

    def save(self):
        """Salva configurações no arquivo."""
        with open(self.config_file, 'w') as f:
            self.config.write(f)

    def get_api_key(self) -> str:
        """Retorna a API key salva."""
        return self.config.get('API', 'gemini_key', fallback='')

    def set_api_key(self, key: str):
        """Salva a API key."""
        if 'API' not in self.config:
            self.config['API'] = {}
        self.config['API']['gemini_key'] = key
        self.save()


class VerificadorGeorreferenciamento:
    """Classe principal da aplicação de verificação de documentos."""

    def __init__(self, root):
        self.root = root
        self.root.title("Verificador INCRA - Sistema Profissional v4.0")
        self.root.geometry("1400x950")

        # Gerenciador de configurações
        self.config_manager = ConfigManager()

        # Variáveis para armazenar caminhos dos arquivos
        self.incra_path = tk.StringVar()
        self.projeto_path = tk.StringVar()
        self.numero_prenotacao = tk.StringVar()
        self.modo_atual = tk.StringVar(value="manual")

        # Variáveis para armazenar dados extraídos
        self.incra_excel_path: Optional[str] = None
        self.projeto_excel_path: Optional[str] = None
        self.incra_data: Optional[Dict] = None
        self.projeto_data: Optional[Dict] = None

        # Variáveis para modo automático
        self.pdf_extraido_incra: Optional[str] = None
        self.pdf_extraido_projeto: Optional[str] = None
        self.preview_incra_image: Optional[Image.Image] = None
        self.preview_projeto_image: Optional[Image.Image] = None

        # Configurar estilo moderno
        self._configurar_estilo()

        # Criar interface
        self._criar_interface()

        # Carregar API key salva
        self._carregar_api_key()

    def _configurar_estilo(self):
        """Configura tema moderno e profissional."""
        style = ttk.Style()
        style.theme_use('clam')

        # Cores modernas
        bg_color = "#f0f0f0"
        accent_color = "#2196F3"
        success_color = "#4CAF50"
        warning_color = "#FF9800"

        # Configurar estilos
        style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'), foreground=accent_color)
        style.configure('Subtitle.TLabel', font=('Segoe UI', 12, 'bold'))
        style.configure('Normal.TLabel', font=('Segoe UI', 10))
        style.configure('Large.TButton', font=('Segoe UI', 11, 'bold'), padding=12)
        style.configure('Action.TButton', font=('Segoe UI', 10, 'bold'), padding=8)
        style.configure('Custom.TNotebook', tabposition='wn')
        style.configure('Custom.TNotebook.Tab', padding=[20, 10], font=('Segoe UI', 11, 'bold'))

        self.root.configure(bg=bg_color)

    def _criar_interface(self):
        """Cria todos os elementos da interface gráfica."""

        # Frame principal com padding
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Configurar grid para expansão
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)

        # ===== CABEÇALHO =====
        header_frame = ttk.Frame(main_frame)
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 15))

        ttk.Label(
            header_frame,
            text="🏛️ VERIFICADOR DE GEORREFERENCIAMENTO INCRA",
            style='Title.TLabel'
        ).pack()

        ttk.Label(
            header_frame,
            text="Sistema Profissional de Análise e Conferência - v4.0",
            style='Normal.TLabel',
            foreground='#666'
        ).pack()

        # ===== BARRA DE FERRAMENTAS SUPERIOR =====
        toolbar_frame = ttk.Frame(main_frame)
        toolbar_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15))

        # Botão de configuração da API
        ttk.Button(
            toolbar_frame,
            text="⚙️ Configurar API Key",
            command=self._abrir_config_api,
            style='Action.TButton'
        ).pack(side=tk.LEFT, padx=5)

        # Indicador de status da API
        self.api_status_label = ttk.Label(
            toolbar_frame,
            text="❌ API Key não configurada",
            foreground='red',
            font=('Segoe UI', 9)
        )
        self.api_status_label.pack(side=tk.LEFT, padx=10)

        ttk.Separator(toolbar_frame, orient='vertical').pack(side=tk.LEFT, fill='y', padx=10)

        # Campo de Número de Prenotação (obrigatório)
        ttk.Label(
            toolbar_frame,
            text="📋 Nº Prenotação:",
            style='Subtitle.TLabel'
        ).pack(side=tk.LEFT, padx=5)

        prenotacao_entry = ttk.Entry(
            toolbar_frame,
            textvariable=self.numero_prenotacao,
            width=15,
            font=('Segoe UI', 11, 'bold')
        )
        prenotacao_entry.pack(side=tk.LEFT, padx=5)

        # Validação para aceitar apenas números
        vcmd = (self.root.register(self._validar_numero), '%P')
        prenotacao_entry.config(validate='key', validatecommand=vcmd)

        ttk.Label(
            toolbar_frame,
            text="(apenas números)",
            font=('Segoe UI', 8),
            foreground='#666'
        ).pack(side=tk.LEFT)

        # ===== NOTEBOOK PARA MODOS =====
        self.notebook = ttk.Notebook(main_frame, style='Custom.TNotebook')
        self.notebook.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # TAB 1: MODO MANUAL
        self.tab_manual = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(self.tab_manual, text="📝 MODO MANUAL")

        # TAB 2: MODO AUTOMÁTICO
        self.tab_automatico = ttk.Frame(self.notebook, padding="20")
        self.notebook.add(self.tab_automatico, text="🤖 MODO AUTOMÁTICO")

        # Criar conteúdo das tabs
        self._criar_modo_manual()
        self._criar_modo_automatico()

        # ===== ÁREA DE RESULTADOS (compartilhada) =====
        result_frame = ttk.LabelFrame(main_frame, text="📊 Relatório de Comparação", padding="10")
        result_frame.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(15, 0))

        main_frame.rowconfigure(3, weight=1)

        # Área de texto com scroll
        self.resultado_text = scrolledtext.ScrolledText(
            result_frame,
            width=120,
            height=12,
            font=('Consolas', 9),
            wrap=tk.WORD
        )
        self.resultado_text.pack(fill=tk.BOTH, expand=True)

        # ===== BARRA DE STATUS =====
        self.status_label = ttk.Label(
            main_frame,
            text="✅ Pronto para iniciar",
            relief=tk.SUNKEN,
            anchor=tk.W,
            font=('Segoe UI', 9)
        )
        self.status_label.grid(row=4, column=0, sticky=(tk.W, tk.E), pady=(10, 0))

    def _criar_modo_manual(self):
        """Cria interface do modo manual."""
        frame = self.tab_manual

        ttk.Label(
            frame,
            text="Selecione manualmente os arquivos para comparação",
            style='Normal.TLabel',
            foreground='#666'
        ).pack(pady=(0, 20))

        # Frame para seleção de arquivos
        files_frame = ttk.Frame(frame)
        files_frame.pack(fill=tk.X, pady=10)

        # INCRA
        incra_frame = ttk.LabelFrame(files_frame, text="📄 Memorial INCRA", padding="15")
        incra_frame.pack(fill=tk.X, pady=10)

        ttk.Entry(
            incra_frame,
            textvariable=self.incra_path,
            font=('Segoe UI', 10),
            state='readonly'
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        ttk.Button(
            incra_frame,
            text="📁 Selecionar PDF",
            command=lambda: self._selecionar_arquivo(self.incra_path, "INCRA"),
            style='Action.TButton'
        ).pack(side=tk.RIGHT)

        # PROJETO
        projeto_frame = ttk.LabelFrame(files_frame, text="📐 Planta/Projeto", padding="15")
        projeto_frame.pack(fill=tk.X, pady=10)

        ttk.Entry(
            projeto_frame,
            textvariable=self.projeto_path,
            font=('Segoe UI', 10),
            state='readonly'
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        ttk.Button(
            projeto_frame,
            text="📁 Selecionar PDF",
            command=lambda: self._selecionar_arquivo(self.projeto_path, "Projeto"),
            style='Action.TButton'
        ).pack(side=tk.RIGHT)

        # Botão de comparação
        ttk.Separator(frame, orient='horizontal').pack(fill=tk.X, pady=20)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack()

        self.btn_comparar_manual = ttk.Button(
            btn_frame,
            text="🔍 COMPARAR DOCUMENTOS",
            command=self._comparar_manual,
            style='Large.TButton',
            width=40
        )
        self.btn_comparar_manual.pack(pady=10)

    def _criar_modo_automatico(self):
        """Cria interface do modo automático."""
        frame = self.tab_automatico

        # Instruções
        info_frame = ttk.Frame(frame)
        info_frame.pack(fill=tk.X, pady=(0, 20))

        ttk.Label(
            info_frame,
            text="🤖 Modo Automático - Busca Inteligente",
            style='Subtitle.TLabel'
        ).pack()

        ttk.Label(
            info_frame,
            text="O sistema buscará automaticamente o arquivo TIFF na rede, converterá para PDF e extrairá os documentos necessários.",
            font=('Segoe UI', 9),
            foreground='#666',
            wraplength=800
        ).pack(pady=5)

        # Campo de entrada para prenotação
        input_frame = ttk.LabelFrame(frame, text="📋 Dados de Entrada", padding="20")
        input_frame.pack(fill=tk.X, pady=20)

        entry_container = ttk.Frame(input_frame)
        entry_container.pack()

        ttk.Label(
            entry_container,
            text="Número de Prenotação:",
            style='Subtitle.TLabel'
        ).pack(side=tk.LEFT, padx=(0, 10))

        self.prenotacao_auto_entry = ttk.Entry(
            entry_container,
            textvariable=self.numero_prenotacao,
            width=20,
            font=('Segoe UI', 12, 'bold')
        )
        self.prenotacao_auto_entry.pack(side=tk.LEFT, padx=5)

        vcmd = (self.root.register(self._validar_numero), '%P')
        self.prenotacao_auto_entry.config(validate='key', validatecommand=vcmd)

        ttk.Label(
            entry_container,
            text="Ex: 229885 (sem zeros à esquerda)",
            font=('Segoe UI', 8, 'italic'),
            foreground='#666'
        ).pack(side=tk.LEFT, padx=10)

        # Botão de iniciar busca automática
        ttk.Separator(frame, orient='horizontal').pack(fill=tk.X, pady=20)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack()

        self.btn_iniciar_automatico = ttk.Button(
            btn_frame,
            text="🚀 INICIAR BUSCA AUTOMÁTICA",
            command=self._iniciar_modo_automatico,
            style='Large.TButton',
            width=40
        )
        self.btn_iniciar_automatico.pack(pady=10)

        # Frame para preview dos documentos (inicialmente oculto)
        self.preview_frame = ttk.LabelFrame(frame, text="👁️ Prévia dos Documentos Extraídos", padding="20")

        preview_container = ttk.Frame(self.preview_frame)
        preview_container.pack(fill=tk.BOTH, expand=True)

        # Preview INCRA
        incra_prev_frame = ttk.Frame(preview_container)
        incra_prev_frame.pack(side=tk.LEFT, padx=20, expand=True)

        ttk.Label(incra_prev_frame, text="📄 Memorial INCRA", style='Subtitle.TLabel').pack(pady=5)
        self.incra_preview_label = ttk.Label(incra_prev_frame, text="", relief=tk.RIDGE)
        self.incra_preview_label.pack()

        # Preview PROJETO
        projeto_prev_frame = ttk.Frame(preview_container)
        projeto_prev_frame.pack(side=tk.LEFT, padx=20, expand=True)

        ttk.Label(projeto_prev_frame, text="📐 Planta/Projeto", style='Subtitle.TLabel').pack(pady=5)
        self.projeto_preview_label = ttk.Label(projeto_prev_frame, text="", relief=tk.RIDGE)
        self.projeto_preview_label.pack()

        # Botões de confirmação
        confirm_frame = ttk.Frame(self.preview_frame)
        confirm_frame.pack(pady=20)

        ttk.Button(
            confirm_frame,
            text="✅ CONTINUAR",
            command=self._confirmar_documentos_automaticos,
            style='Large.TButton',
            width=25
        ).pack(side=tk.LEFT, padx=10)

        ttk.Button(
            confirm_frame,
            text="✋ FAZER MANUAL",
            command=self._alternar_para_manual,
            style='Large.TButton',
            width=25
        ).pack(side=tk.LEFT, padx=10)

    def _validar_numero(self, valor):
        """Valida entrada para aceitar apenas números."""
        return valor == "" or valor.isdigit()

    def _carregar_api_key(self):
        """Carrega API key salva e atualiza interface."""
        api_key = self.config_manager.get_api_key()
        if api_key:
            self.api_status_label.config(
                text="✅ API Key configurada",
                foreground='green'
            )
        else:
            self.api_status_label.config(
                text="❌ API Key não configurada",
                foreground='red'
            )

    def _abrir_config_api(self):
        """Abre janela para configurar API key."""
        config_window = tk.Toplevel(self.root)
        config_window.title("Configuração da API Key")
        config_window.geometry("600x250")
        config_window.transient(self.root)
        config_window.grab_set()

        main_frame = ttk.Frame(config_window, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(
            main_frame,
            text="🔑 Configuração da API Key do Gemini",
            style='Subtitle.TLabel'
        ).pack(pady=(0, 10))

        ttk.Label(
            main_frame,
            text="Insira sua API key abaixo. Ela será salva de forma segura e não precisará ser inserida novamente.",
            wraplength=500,
            font=('Segoe UI', 9)
        ).pack(pady=10)

        # Campo de entrada
        api_var = tk.StringVar(value=self.config_manager.get_api_key())

        entry_frame = ttk.Frame(main_frame)
        entry_frame.pack(fill=tk.X, pady=20)

        ttk.Label(entry_frame, text="API Key:", font=('Segoe UI', 10, 'bold')).pack(anchor=tk.W)

        api_entry = ttk.Entry(
            entry_frame,
            textvariable=api_var,
            font=('Segoe UI', 10),
            show="*",
            width=60
        )
        api_entry.pack(fill=tk.X, pady=5)

        # Botões
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=20)

        def salvar_api():
            key = api_var.get().strip()
            if key:
                self.config_manager.set_api_key(key)
                self._carregar_api_key()
                messagebox.showinfo("Sucesso", "API Key salva com sucesso!")
                config_window.destroy()
            else:
                messagebox.showwarning("Aviso", "Por favor, insira uma API Key válida.")

        ttk.Button(
            btn_frame,
            text="💾 Salvar",
            command=salvar_api,
            style='Action.TButton'
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            btn_frame,
            text="❌ Cancelar",
            command=config_window.destroy,
            style='Action.TButton'
        ).pack(side=tk.LEFT, padx=5)

    def _selecionar_arquivo(self, variavel, tipo):
        """Abre diálogo para selecionar arquivo PDF."""
        filename = filedialog.askopenfilename(
            title=f"Selecionar arquivo {tipo}",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")]
        )
        if filename:
            variavel.set(filename)

    def _atualizar_status(self, mensagem: str, cor: str = 'black'):
        """Atualiza a barra de status."""
        self.status_label.config(text=mensagem)
        self.root.update_idletasks()

    def _desabilitar_botoes(self):
        """Desabilita botões durante o processamento."""
        self.btn_comparar_manual.config(state='disabled')
        self.btn_iniciar_automatico.config(state='disabled')

    def _habilitar_botoes(self):
        """Reabilita botões após o processamento."""
        self.btn_comparar_manual.config(state='normal')
        self.btn_iniciar_automatico.config(state='normal')

    # ========== MODO MANUAL ==========

    def _comparar_manual(self):
        """Executa comparação no modo manual."""
        if not self._validar_entrada_manual():
            return

        def executar():
            try:
                self._desabilitar_botoes()
                self._atualizar_status("🔄 Processando documentos...")

                # Extrair dados para Excel
                self._atualizar_status("📄 Extraindo dados do INCRA...")
                self.incra_excel_path, self.incra_data = self._extrair_pdf_para_excel(
                    self.incra_path.get(), "incra"
                )

                self._atualizar_status("📐 Extraindo dados do Projeto...")
                self.projeto_excel_path, self.projeto_data = self._extrair_pdf_para_excel(
                    self.projeto_path.get(), "normal"
                )

                # Gerar relatório
                self._atualizar_status("📊 Gerando relatório de comparação...")
                relatorio = self._construir_relatorio_comparacao(True, False)

                # Salvar e abrir relatório
                self._salvar_e_abrir_relatorio(relatorio)

                # Mostrar resumo
                self._mostrar_resumo_no_texto()

                self._atualizar_status("✅ Comparação concluída com sucesso!")

            except Exception as e:
                self._atualizar_status(f"❌ Erro: {str(e)}")
                messagebox.showerror("Erro", f"Erro ao processar documentos:\n\n{str(e)}")
            finally:
                self._habilitar_botoes()

        # Executar em thread separada
        threading.Thread(target=executar, daemon=True).start()

    def _validar_entrada_manual(self) -> bool:
        """Valida entradas do modo manual."""
        api_key = self.config_manager.get_api_key()
        if not api_key:
            messagebox.showerror("Erro", "Por favor, configure a API Key primeiro.")
            return False

        if not self.incra_path.get():
            messagebox.showerror("Erro", "Por favor, selecione o arquivo INCRA.")
            return False

        if not self.projeto_path.get():
            messagebox.showerror("Erro", "Por favor, selecione o arquivo Projeto/Planta.")
            return False

        if not self.numero_prenotacao.get():
            messagebox.showerror("Erro", "Por favor, insira o Número de Prenotação.")
            return False

        return True

    # ========== MODO AUTOMÁTICO ==========

    def _iniciar_modo_automatico(self):
        """Inicia o processo automático."""
        if not self._validar_entrada_automatico():
            return

        def executar():
            try:
                self._desabilitar_botoes()

                # 1. Buscar arquivo TIFF
                self._atualizar_status("🔍 Buscando arquivo TIFF na rede...")
                tiff_path = self._buscar_arquivo_tiff()

                if not tiff_path:
                    raise Exception("Arquivo TIFF não encontrado na rede.")

                # 2. Copiar e converter para PDF
                self._atualizar_status("📋 Copiando e convertendo TIFF para PDF...")
                pdf_path = self._converter_tiff_para_pdf(tiff_path)

                # 3. Extrair documentos do PDF
                self._atualizar_status("📄 Extraindo Memorial INCRA...")
                self.pdf_extraido_incra = self._extrair_memorial_incra_do_pdf(pdf_path)

                self._atualizar_status("📐 Extraindo Planta/Projeto...")
                self.pdf_extraido_projeto = self._extrair_projeto_do_pdf(pdf_path)

                # 4. Salvar backups
                self._atualizar_status("💾 Salvando backups...")
                self._salvar_backups_pdfs()

                # 5. Gerar previews
                self._atualizar_status("👁️ Gerando prévias...")
                self._gerar_previews()

                # 6. Mostrar frame de preview
                self.preview_frame.pack(fill=tk.BOTH, expand=True, pady=20)

                self._atualizar_status("✅ Documentos extraídos! Verifique as prévias.")

            except Exception as e:
                self._atualizar_status(f"❌ Erro: {str(e)}")
                messagebox.showerror("Erro", f"Erro no modo automático:\n\n{str(e)}")
                self._habilitar_botoes()

        # Executar em thread separada
        threading.Thread(target=executar, daemon=True).start()

    def _validar_entrada_automatico(self) -> bool:
        """Valida entradas do modo automático."""
        api_key = self.config_manager.get_api_key()
        if not api_key:
            messagebox.showerror("Erro", "Por favor, configure a API Key primeiro.")
            return False

        if not self.numero_prenotacao.get():
            messagebox.showerror("Erro", "Por favor, insira o Número de Prenotação.")
            return False

        return True

    def _buscar_arquivo_tiff(self) -> Optional[str]:
        """Busca arquivo TIFF na rede baseado no número de prenotação."""
        # Obter número e formatar
        numero = int(self.numero_prenotacao.get())
        numero_formatado = f"{numero:08d}"  # 8 dígitos com zeros à esquerda

        # Calcular subpasta (milhar superior)
        milhar = math.ceil(numero / 1000) * 1000
        subpasta_formatada = f"{milhar:08d}"

        # Montar caminho
        base_path = Path(r"\\192.168.20.100\trabalho\TRABALHO\IMAGENS\IMOVEIS\DOCUMENTOS - DIVERSOS")
        tiff_path = base_path / subpasta_formatada / f"{numero_formatado}.tif"

        self._atualizar_status(f"🔍 Buscando: {tiff_path}")

        if tiff_path.exists():
            return str(tiff_path)

        return None

    def _converter_tiff_para_pdf(self, tiff_path: str) -> str:
        """Copia TIFF para Downloads e converte para PDF multi-página."""
        downloads_dir = Path.home() / "Downloads" / "conferencia_geo_temp"
        downloads_dir.mkdir(parents=True, exist_ok=True)

        # Copiar TIFF
        tiff_filename = Path(tiff_path).name
        tiff_dest = downloads_dir / tiff_filename
        shutil.copy2(tiff_path, tiff_dest)

        # Converter para PDF
        pdf_path = downloads_dir / f"{Path(tiff_filename).stem}.pdf"

        # Abrir TIFF multi-página
        img = Image.open(tiff_dest)
        images = []

        try:
            while True:
                images.append(img.copy().convert('RGB'))
                img.seek(img.tell() + 1)
        except EOFError:
            pass  # Fim das páginas

        # Salvar como PDF
        if images:
            images[0].save(
                pdf_path,
                save_all=True,
                append_images=images[1:],
                resolution=200.0
            )

        return str(pdf_path)

    def _extrair_memorial_incra_do_pdf(self, pdf_path: str) -> str:
        """Extrai páginas do Memorial INCRA do PDF."""
        output_dir = Path.home() / "Downloads" / "conferencia_geo_temp"
        output_pdf = output_dir / "memorial_incra_extraido.pdf"

        # Usar Gemini para identificar as páginas relevantes
        api_key = self.config_manager.get_api_key()
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-1.5-flash')

        # Converter PDF para imagens
        images = convert_from_path(pdf_path, dpi=150)

        paginas_encontradas = []

        for i, img in enumerate(images):
            # Salvar imagem temporária
            temp_img_path = output_dir / f"temp_page_{i}.jpg"
            img.save(temp_img_path, 'JPEG')

            # Analisar com Gemini
            prompt = """
            Analise esta imagem e responda apenas com 'SIM' ou 'NAO':
            Esta página contém o Memorial Descritivo do INCRA?

            Características do Memorial INCRA:
            - Texto: "MINISTÉRIO DA AGRICULTURA, PECUÁRIA E ABASTECIMENTO"
            - Texto: "INSTITUTO NACIONAL DE COLONIZAÇÃO E REFORMA AGRÁRIA"
            - Texto: "MEMORIAL DESCRITIVO"
            - Tabela com colunas: "VÉRTICE", "SEGMENTO VANTE", "Confrontações"
            - Texto: "DESCRIÇÃO DA PARCELA"

            Responda apenas: SIM ou NAO
            """

            try:
                img_upload = Image.open(temp_img_path)
                response = model.generate_content([prompt, img_upload])
                resposta = response.text.strip().upper()

                if 'SIM' in resposta:
                    paginas_encontradas.append(i)

            except Exception as e:
                print(f"Erro ao analisar página {i}: {e}")

            # Limpar imagem temporária
            temp_img_path.unlink()

        # Extrair páginas encontradas para novo PDF
        if paginas_encontradas:
            with open(pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                writer = PyPDF2.PdfWriter()

                for page_num in paginas_encontradas:
                    writer.add_page(reader.pages[page_num])

                with open(output_pdf, 'wb') as output_file:
                    writer.write(output_file)

        return str(output_pdf)

    def _extrair_projeto_do_pdf(self, pdf_path: str) -> str:
        """Extrai páginas da Planta/Projeto do PDF."""
        output_dir = Path.home() / "Downloads" / "conferencia_geo_temp"
        output_pdf = output_dir / "projeto_extraido.pdf"

        # Usar Gemini para identificar as páginas relevantes
        api_key = self.config_manager.get_api_key()
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-1.5-flash')

        # Converter PDF para imagens
        images = convert_from_path(pdf_path, dpi=150)

        paginas_encontradas = []

        for i, img in enumerate(images):
            # Salvar imagem temporária
            temp_img_path = output_dir / f"temp_page_{i}.jpg"
            img.save(temp_img_path, 'JPEG')

            # Analisar com Gemini
            prompt = """
            Analise esta imagem e responda apenas com 'SIM' ou 'NAO':
            Esta página contém a Planta/Projeto de Georreferenciamento?

            Características da Planta/Projeto:
            - Títulos: "PLANTA DO IMÓVEL GEORREFERENCIADO" ou "PLANTA DE SITUAÇÃO"
            - Identificadores: "Código INCRA:", "Matrícula nº:", "Responsável técnico:", "Propriedade:", "Município:"
            - Tabela com coordenadas (colunas: "Código", "Longitude", "Latitude")
            - Desenho/mapa com vértices (ex: AKE-M-1028)

            Responda apenas: SIM ou NAO
            """

            try:
                img_upload = Image.open(temp_img_path)
                response = model.generate_content([prompt, img_upload])
                resposta = response.text.strip().upper()

                if 'SIM' in resposta:
                    paginas_encontradas.append(i)

            except Exception as e:
                print(f"Erro ao analisar página {i}: {e}")

            # Limpar imagem temporária
            temp_img_path.unlink()

        # Extrair páginas encontradas para novo PDF
        if paginas_encontradas:
            with open(pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                writer = PyPDF2.PdfWriter()

                for page_num in paginas_encontradas:
                    writer.add_page(reader.pages[page_num])

                with open(output_pdf, 'wb') as output_file:
                    writer.write(output_file)

        return str(output_pdf)

    def _salvar_backups_pdfs(self):
        """Salva backups dos PDFs extraídos."""
        docs_dir = Path.home() / "Documentos" / "Relatórios INCRA"

        # Criar diretórios
        incra_dir = docs_dir / "PDF_INCRAS"
        projeto_dir = docs_dir / "PDF_PLANTAS"

        incra_dir.mkdir(parents=True, exist_ok=True)
        projeto_dir.mkdir(parents=True, exist_ok=True)

        # Nome baseado na prenotação
        numero = self.numero_prenotacao.get()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Copiar PDFs
        if self.pdf_extraido_incra:
            dest_incra = incra_dir / f"INCRA_{numero}_{timestamp}.pdf"
            shutil.copy2(self.pdf_extraido_incra, dest_incra)

        if self.pdf_extraido_projeto:
            dest_projeto = projeto_dir / f"PROJETO_{numero}_{timestamp}.pdf"
            shutil.copy2(self.pdf_extraido_projeto, dest_projeto)

    def _gerar_previews(self):
        """Gera thumbnails dos documentos extraídos."""
        # Preview INCRA
        if self.pdf_extraido_incra and Path(self.pdf_extraido_incra).exists():
            images = convert_from_path(self.pdf_extraido_incra, dpi=100, first_page=1, last_page=1)
            if images:
                self.preview_incra_image = images[0]
                # Redimensionar para thumbnail
                self.preview_incra_image.thumbnail((300, 400))

                photo = ImageTk.PhotoImage(self.preview_incra_image)
                self.incra_preview_label.config(image=photo)
                self.incra_preview_label.image = photo  # Manter referência

        # Preview PROJETO
        if self.pdf_extraido_projeto and Path(self.pdf_extraido_projeto).exists():
            images = convert_from_path(self.pdf_extraido_projeto, dpi=100, first_page=1, last_page=1)
            if images:
                self.preview_projeto_image = images[0]
                # Redimensionar para thumbnail
                self.preview_projeto_image.thumbnail((300, 400))

                photo = ImageTk.PhotoImage(self.preview_projeto_image)
                self.projeto_preview_label.config(image=photo)
                self.projeto_preview_label.image = photo  # Manter referência

    def _confirmar_documentos_automaticos(self):
        """Usuário confirmou documentos - prosseguir com comparação."""
        # Usar os PDFs extraídos como arquivos de entrada
        self.incra_path.set(self.pdf_extraido_incra)
        self.projeto_path.set(self.pdf_extraido_projeto)

        # Ocultar preview
        self.preview_frame.pack_forget()

        # Executar comparação
        self._comparar_manual()

    def _alternar_para_manual(self):
        """Usuário optou por fazer manual - alternar para aba manual."""
        self.preview_frame.pack_forget()
        self.notebook.select(self.tab_manual)
        self._habilitar_botoes()
        messagebox.showinfo(
            "Modo Manual",
            "Selecione manualmente os arquivos corretos na aba 'Modo Manual'."
        )

    # ========== EXTRAÇÃO E COMPARAÇÃO ==========

    def _extrair_pdf_para_excel(self, pdf_path: str, tipo: str = "normal") -> tuple[str, Dict]:
        """
        Extrai dados de um PDF memorial para Excel usando Gemini API.
        """
        try:
            api_key = self.config_manager.get_api_key()
            genai.configure(api_key=api_key)

            # Criar diretório temporário
            output_dir = Path(tempfile.gettempdir()) / "conferencia_geo"
            output_dir.mkdir(parents=True, exist_ok=True)

            if not output_dir.exists():
                raise RuntimeError(f"Não foi possível criar o diretório: {output_dir}")

            # Nome do arquivo Excel
            nome_base = Path(pdf_path).stem
            excel_path = output_dir / f"{nome_base}_extraido.xlsx"

            # Extrair dados usando funções existentes
            if tipo == "incra":
                dados = extrair_memorial_incra(pdf_path, api_key)
            else:
                dados = extract_table_from_pdf(pdf_path, api_key)

            if not dados or 'data' not in dados:
                raise ValueError("Nenhum dado foi extraído do PDF")

            # Criar arquivo Excel
            create_excel_file(dados, str(excel_path))

            if not excel_path.exists():
                raise RuntimeError(f"Arquivo Excel não foi criado")

            if excel_path.stat().st_size == 0:
                raise RuntimeError(f"Arquivo Excel está vazio")

            return str(excel_path), dados

        except Exception as e:
            error_msg = f"Erro ao extrair PDF para Excel: {str(e)}"
            raise RuntimeError(error_msg) from e

    def _normalizar_coordenada(self, coord: str) -> str:
        """
        Normaliza coordenadas para comparação, ignorando diferenças de formato.
        """
        if not coord:
            return ""

        coord = str(coord).strip()

        # Normalizar caracteres Unicode especiais
        coord = coord.replace("′", "'").replace("″", '"')

        # Remover "-" do início (INCRA)
        if coord.startswith("-"):
            coord = coord[1:].strip()

        # Remover " W" ou " S" do final (PROJETO)
        coord = coord.replace(" W", "").replace(" S", "").strip()

        # Remover aspas e espaços extras
        coord = coord.strip().strip('"').strip("'").strip()

        return coord

    def _limpar_string(self, valor) -> str:
        """
        Limpa qualquer valor convertendo para string e removendo espaços em branco.
        Converte pontos decimais em vírgulas para padronização numérica brasileira.
        """
        if valor is None:
            return ""

        valor_limpo = str(valor).strip()

        # Remover espaços duplos internos
        while "  " in valor_limpo:
            valor_limpo = valor_limpo.replace("  ", " ")

        # Converter ponto decimal para vírgula (padrão brasileiro)
        valor_limpo = valor_limpo.replace(".", ",")

        return valor_limpo

    def _construir_relatorio_comparacao(self, incluir_projeto: bool, incluir_memorial: bool) -> str:
        """
        Constrói relatório HTML comparando dados estruturados.
        """
        html = []

        # Cabeçalho HTML
        html.append("""<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Relatório de Conferência INCRA</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
        }
        .container {
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            padding: 40px;
            border-radius: 15px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
        }
        h1 {
            color: #2c3e50;
            text-align: center;
            margin-bottom: 10px;
            font-size: 32px;
        }
        .subtitle {
            text-align: center;
            color: #7f8c8d;
            margin-bottom: 30px;
            font-size: 14px;
        }
        .info-box {
            background: #ecf0f1;
            padding: 15px;
            border-radius: 8px;
            margin-bottom: 30px;
        }
        .info-box strong {
            color: #2c3e50;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 30px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        th {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 15px;
            text-align: left;
            font-weight: 600;
            font-size: 14px;
        }
        td {
            padding: 12px 15px;
            border-bottom: 1px solid #ecf0f1;
            font-size: 13px;
        }
        tr:hover {
            background-color: #f8f9fa;
        }
        .identico {
            background-color: #d4edda !important;
            border-left: 4px solid #28a745;
        }
        .diferente {
            background-color: #f8d7da !important;
            border-left: 4px solid #dc3545;
            font-weight: 600;
        }
        .resumo {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 25px;
            border-radius: 10px;
            margin-top: 30px;
        }
        .resumo h2 {
            margin-bottom: 20px;
            font-size: 24px;
        }
        .resumo h4 {
            margin-top: 15px;
            margin-bottom: 10px;
            font-size: 18px;
        }
        .resumo p {
            margin: 5px 0;
            font-size: 16px;
        }
        .section-title {
            color: #2c3e50;
            margin: 40px 0 20px 0;
            padding-bottom: 10px;
            border-bottom: 3px solid #667eea;
            font-size: 24px;
        }
        .rodape {
            text-align: center;
            margin-top: 40px;
            padding-top: 20px;
            border-top: 2px solid #ecf0f1;
            color: #7f8c8d;
            font-size: 12px;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>📋 RELATÓRIO DE CONFERÊNCIA INCRA</h1>
        <p class="subtitle">Sistema Profissional de Análise e Verificação v4.0</p>
""")

        # Informações do relatório
        html.append(f"""
        <div class="info-box">
            <p><strong>📅 Data:</strong> {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}</p>
            <p><strong>📋 Nº Prenotação:</strong> {self.numero_prenotacao.get()}</p>
        </div>
""")

        # Carregar dados dos Excel
        wb_incra = load_workbook(self.incra_excel_path)
        ws_incra = wb_incra.active
        dados_incra = list(ws_incra.iter_rows(values_only=True))

        wb_projeto = load_workbook(self.projeto_excel_path)
        ws_projeto = wb_projeto.active
        dados_projeto = list(ws_projeto.iter_rows(values_only=True))

        # Contadores
        identicos_vertice = 0
        diferencas_vertice = 0
        identicos_segmento = 0
        diferencas_segmento = 0

        # ===== SEÇÃO 1: VÉRTICE =====
        html.append('<h2 class="section-title">📐 COMPARAÇÃO: VÉRTICE (Código, Longitude, Latitude, Altitude)</h2>')
        html.append('<table>')
        html.append('<thead><tr>')
        html.append('<th>Linha</th>')
        html.append('<th>Campo</th>')
        html.append('<th>INCRA</th>')
        html.append('<th>PROJETO</th>')
        html.append('<th>Status</th>')
        html.append('</tr></thead>')
        html.append('<tbody>')

        max_rows = max(len(dados_incra), len(dados_projeto))

        for i in range(1, max_rows):  # Pular cabeçalho
            incra_row = dados_incra[i] if i < len(dados_incra) else []
            projeto_row = dados_projeto[i] if i < len(dados_projeto) else []

            # Extrair e limpar dados (colunas 0-3: Código, Longitude, Latitude, Altitude)
            codigo_incra = self._limpar_string(incra_row[0] if len(incra_row) > 0 else "")
            codigo_projeto = self._limpar_string(projeto_row[0] if len(projeto_row) > 0 else "")

            long_incra = self._normalizar_coordenada(self._limpar_string(incra_row[1] if len(incra_row) > 1 else ""))
            long_projeto = self._normalizar_coordenada(self._limpar_string(projeto_row[1] if len(projeto_row) > 1 else ""))

            lat_incra = self._normalizar_coordenada(self._limpar_string(incra_row[2] if len(incra_row) > 2 else ""))
            lat_projeto = self._normalizar_coordenada(self._limpar_string(projeto_row[2] if len(projeto_row) > 2 else ""))

            alt_incra = self._limpar_string(incra_row[3] if len(incra_row) > 3 else "")
            alt_projeto = self._limpar_string(projeto_row[3] if len(projeto_row) > 3 else "")

            # Comparar cada campo
            campos = [
                ("Código", codigo_incra, codigo_projeto),
                ("Longitude", long_incra, long_projeto),
                ("Latitude", lat_incra, lat_projeto),
                ("Altitude", alt_incra, alt_projeto)
            ]

            for campo, val_incra, val_projeto in campos:
                status_classe = "identico" if val_incra == val_projeto else "diferente"
                status_texto = "✅ Idêntico" if val_incra == val_projeto else "❌ Diferente"

                if val_incra == val_projeto:
                    identicos_vertice += 1
                else:
                    diferencas_vertice += 1

                html.append(f'<tr class="{status_classe}">')
                html.append(f'<td>{i}</td>')
                html.append(f'<td><strong>{campo}</strong></td>')
                html.append(f'<td>{val_incra}</td>')
                html.append(f'<td>{val_projeto}</td>')
                html.append(f'<td>{status_texto}</td>')
                html.append('</tr>')

        html.append('</tbody></table>')

        # ===== SEÇÃO 2: SEGMENTO VANTE =====
        html.append('<h2 class="section-title">🔄 COMPARAÇÃO: SEGMENTO VANTE (Código, Azimute, Distância)</h2>')
        html.append('<table>')
        html.append('<thead><tr>')
        html.append('<th>Linha</th>')
        html.append('<th>Campo</th>')
        html.append('<th>INCRA</th>')
        html.append('<th>PROJETO</th>')
        html.append('<th>Status</th>')
        html.append('</tr></thead>')
        html.append('<tbody>')

        for i in range(1, max_rows):
            incra_row = dados_incra[i] if i < len(dados_incra) else []
            projeto_row = dados_projeto[i] if i < len(dados_projeto) else []

            # Extrair e limpar dados (colunas 4-6: Código, Azimute, Distância)
            cod_seg_incra = self._limpar_string(incra_row[4] if len(incra_row) > 4 else "")
            cod_seg_projeto = self._limpar_string(projeto_row[4] if len(projeto_row) > 4 else "")

            azim_incra = self._limpar_string(incra_row[5] if len(incra_row) > 5 else "")
            azim_projeto = self._limpar_string(projeto_row[5] if len(projeto_row) > 5 else "")

            dist_incra = self._limpar_string(incra_row[6] if len(incra_row) > 6 else "")
            dist_projeto = self._limpar_string(projeto_row[6] if len(projeto_row) > 6 else "")

            # Comparar cada campo
            campos = [
                ("Código", cod_seg_incra, cod_seg_projeto),
                ("Azimute", azim_incra, azim_projeto),
                ("Distância", dist_incra, dist_projeto)
            ]

            for campo, val_incra, val_projeto in campos:
                status_classe = "identico" if val_incra == val_projeto else "diferente"
                status_texto = "✅ Idêntico" if val_incra == val_projeto else "❌ Diferente"

                if val_incra == val_projeto:
                    identicos_segmento += 1
                else:
                    diferencas_segmento += 1

                html.append(f'<tr class="{status_classe}">')
                html.append(f'<td>{i}</td>')
                html.append(f'<td><strong>{campo}</strong></td>')
                html.append(f'<td>{val_incra}</td>')
                html.append(f'<td>{val_projeto}</td>')
                html.append(f'<td>{status_texto}</td>')
                html.append('</tr>')

        html.append('</tbody></table>')

        # ===== RESUMO =====
        identicos_total = identicos_vertice + identicos_segmento
        diferencas_total = diferencas_vertice + diferencas_segmento

        html.append(f"""
        <div class="resumo">
            <h2>📊 RESUMO DA COMPARAÇÃO</h2>

            <h4>📍 VÉRTICE (Código, Longitude, Latitude, Altitude):</h4>
            <p>✅ Idênticos: <strong>{identicos_vertice}</strong></p>
            <p>❌ Diferentes: <strong>{diferencas_vertice}</strong></p>

            <h4>🔄 SEGMENTO VANTE (Código, Azimute, Distância):</h4>
            <p>✅ Idênticos: <strong>{identicos_segmento}</strong></p>
            <p>❌ Diferentes: <strong>{diferencas_segmento}</strong></p>

            <h4>🎯 TOTAL GERAL:</h4>
            <p>✅ Total idênticos: <strong>{identicos_total}</strong></p>
            <p>❌ Total diferentes: <strong>{diferencas_total}</strong></p>
        </div>
""")

        # Rodapé
        html.append("""
        <div class="rodape">
            <p>Relatório gerado automaticamente pelo Sistema de Verificação INCRA v4.0</p>
            <p>© 2024 - Todos os direitos reservados</p>
        </div>
    </div>
</body>
</html>
""")

        return "".join(html)

    def _salvar_e_abrir_relatorio(self, conteudo_html: str):
        """Salva relatório automaticamente e abre no navegador."""
        # Criar diretório se não existir
        relatorios_dir = Path.home() / "Documentos" / "Relatórios INCRA"
        relatorios_dir.mkdir(parents=True, exist_ok=True)

        # Nome do arquivo
        numero = self.numero_prenotacao.get()
        nome_arquivo = f"Relatório_INCRA_{numero}.html"
        caminho_completo = relatorios_dir / nome_arquivo

        # Salvar arquivo
        with open(caminho_completo, 'w', encoding='utf-8') as f:
            f.write(conteudo_html)

        # Abrir no navegador
        webbrowser.open(f'file://{caminho_completo}')

        self._atualizar_status(f"✅ Relatório salvo: {caminho_completo}")

    def _mostrar_resumo_no_texto(self):
        """Mostra resumo simplificado na área de texto."""
        self.resultado_text.delete(1.0, tk.END)

        resumo = f"""
╔════════════════════════════════════════════════════════════════╗
║          COMPARAÇÃO CONCLUÍDA COM SUCESSO                      ║
╚════════════════════════════════════════════════════════════════╝

📋 Número de Prenotação: {self.numero_prenotacao.get()}
📅 Data: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}

✅ O relatório HTML completo foi gerado e aberto automaticamente.
📁 Local: Documentos\\Relatórios INCRA\\Relatório_INCRA_{self.numero_prenotacao.get()}.html

💡 Consulte o relatório HTML para ver todos os detalhes da comparação.
"""

        self.resultado_text.insert(1.0, resumo)


def main():
    """Função principal."""
    root = tk.Tk()
    app = VerificadorGeorreferenciamento(root)
    root.mainloop()


if __name__ == "__main__":
    main()
